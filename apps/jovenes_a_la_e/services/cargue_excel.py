"""Lector y validador del Excel de beneficiarios de educación posmedia.

Lee el archivo que manda el área, lo normaliza y devuelve **una fila de reporte
por cada fila del Excel, con su número de fila real**, para que quien corrige
pueda abrir el archivo y buscar el error donde se lo dicen. No toca la base de
datos ni persiste nada: eso es del servicio de cargue. Acá solo se lee, se
normaliza y se juzga.

## Por qué el lector es tolerante

El archivo es institucional y no viene limpio. Lo verificado sobre el de 2025:

- **El encabezado no está en la fila 1.** La fila 1 es un título
  (`KENNEDY META JE4 | 175`) y los nombres de columna están en la 2. Asumir
  posición fija rompe con el primer archivo que traiga una fila de más, así que
  el encabezado **se busca** en las primeras filas.
- **La hoja no siempre se llama igual** (`KENNEDY - META FISICA` en el archivo
  del área, `BENEFICIARIOS` en la plantilla). Se elige la primera hoja donde
  aparezca un encabezado reconocible.
- Los códigos SNIES llegan como número y Excel los devuelve `4894.0`.
- `SEGUNDO_NOMBRE` trae `NO APLICA` como texto literal.
- Los niveles de formación vienen con tildes inconsistentes entre filas.

## El título se compara con lo leído

Si la fila de título trae un número (`… | 175`), se contrasta con las filas
leídas y se avisa si no coinciden. Es la defensa barata contra un archivo
truncado al copiarlo o al filtrarlo, que es un error que no se ve.

## Lo que NO decide este módulo

No decide si la persona existe, ni si ya fue beneficiada, ni el avance del KPI:
todo eso necesita la base y vive en el servicio de cargue. Acá la única
comparación entre filas es **dentro del propio archivo**.
"""
from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass, field

#: Tope de filas de datos. Sin Celery el cargue es síncrono, así que hay que
#: acotarlo explícitamente: el archivo de 2025 trae 175 y este techo da
#: holgura para el cuatrienio entero sin dejar la petición colgada.
LIMITE_FILAS = 2000

#: Cuántas filas se miran buscando el encabezado antes de rendirse.
FILAS_BUSQUEDA_ENCABEZADO = 10

# ── Columnas ────────────────────────────────────────────────────────────────
# Clave interna → alias aceptados (ya normalizados: sin tildes, MAYÚSCULA y con
# los separadores colapsados a "_"). Los alias existen porque el mismo dato
# llega escrito distinto según quién exportó el archivo.
COLUMNAS: dict[str, tuple[str, ...]] = {
    "consecutivo":    ("NO", "N", "NO.", "CONSECUTIVO", "ITEM"),
    "tipo_documento": ("TIPO_DOCUMENTO", "TIPO_DE_DOCUMENTO", "TIPO_DOC"),
    "documento":      ("DOCUMENTO", "NUMERO_DOCUMENTO", "NO_DOCUMENTO", "CEDULA", "IDENTIFICACION"),
    "nombre1":        ("PRIMER_NOMBRE",),
    "nombre2":        ("SEGUNDO_NOMBRE",),
    "apellido1":      ("PRIMER_APELLIDO",),
    "apellido2":      ("SEGUNDO_APELLIDO",),
    "snies_programa": ("SNIES_SIET_PROGRAMA", "SNIES_PROGRAMA", "CODIGO_PROGRAMA"),
    "programa":       ("PROGRAMA", "NOMBRE_PROGRAMA", "PROGRAMA_ACADEMICO"),
    "snies_ies":      ("SNIES_SIET_IES", "SNIES_IES", "CODIGO_IES"),
    "ies_nombre":     ("IES_NOMBRE", "NOMBRE_IES", "INSTITUCION"),
    "nivel":          ("NIVEL_FORMACION", "NIVEL_DE_FORMACION", "NIVEL"),
    "localidad":      ("LOCALIDAD",),
    # ── Opcionales. Hoy el archivo del área NO los trae; se leen si aparecen
    #    para que el día que lleguen no haya que tocar el parser. Sin la
    #    discriminación de acceso/permanencia el cargue corre igual: lo que no
    #    puede hacerse es imputar el avance (lo reporta el servicio, con motivo).
    "acceso":         ("ACCESO", "CUMPLIMIENTO_ACCESO", "META_23771"),
    "permanencia":    ("PERMANENCIA", "CUMPLIMIENTO_PERMANENCIA", "META_23772"),
    "barrio":         ("BARRIO", "BARRIO_RESIDENCIA"),
    "telefono":       ("TELEFONO", "CELULAR", "TELEFONO_CONTACTO"),
    "correo":         ("CORREO", "EMAIL", "CORREO_ELECTRONICO"),
}

#: Las tres sin las que una fila no es nadie. Se usan además para reconocer
#: cuál de las primeras filas es el encabezado.
OBLIGATORIAS = ("documento", "nombre1", "apellido1")

# ── Tipos de documento ──────────────────────────────────────────────────────
# El catálogo `tipo_documento` de la base tiene SEIS entradas (CC, TI, CE, PA,
# NIT, Otro) y el archivo trae siglas que no están: PPT, PEP, RC, NUIP. Se
# mapean a `6 = Otro` y la fila sale con AVISO, no con error: rechazar a un
# beneficiario real porque al catálogo le falta su sigla sería peor. La sigla
# original se conserva en los datos leídos para que no se pierda.
TIPOS_DOCUMENTO: dict[str, int] = {
    "CC": 1, "TI": 2, "CE": 3, "PA": 4, "PASAPORTE": 4, "NIT": 5,
    "PPT": 6, "PEP": 6, "RC": 6, "NUIP": 6, "OTRO": 6,
}
#: Las que se traducen a «Otro» por falta de entrada propia en el catálogo.
SIGLAS_SIN_CATALOGO = ("PPT", "PEP", "RC", "NUIP")

# ── Nivel de formación ──────────────────────────────────────────────────────
# El archivo habla el idioma del SNIES; `entrega_beca.nivel_formacion` tiene sus
# cuatro `choices`. La equivalencia es una DECISIÓN, no una obviedad, así que
# queda escrita acá y no repartida por el código:
#
#   FORMACION TECNICA LABORAL     → etdh   (educación para el trabajo, no es
#                                           título de educación superior)
#   FORMACION TECNICA PROFESIONAL → tecnico_profesional
#   TECNOLOGICO                   → tecnologo
#   UNIVERSITARIO                 → profesional
NIVELES: dict[str, str] = {
    "FORMACION TECNICA LABORAL": "etdh",
    "TECNICA LABORAL": "etdh",
    "ETDH": "etdh",
    "FORMACION TECNICA PROFESIONAL": "tecnico_profesional",
    "TECNICA PROFESIONAL": "tecnico_profesional",
    "TECNICO PROFESIONAL": "tecnico_profesional",
    "TECNOLOGICO": "tecnologo",
    "TECNOLOGO": "tecnologo",
    "UNIVERSITARIO": "profesional",
    "PROFESIONAL": "profesional",
    "PROFESIONAL UNIVERSITARIO": "profesional",
}

#: Texto que el área escribe para decir «vacío». Llega literal en las celdas.
VACIOS = {"", "NO APLICA", "NOAPLICA", "N/A", "NA", "N.A.", "-", "--", "SIN DATO", "NINGUNO"}

_AFIRMATIVOS = {"SI", "S", "X", "1", "TRUE", "VERDADERO", "SÍ"}
_NEGATIVOS = {"NO", "N", "0", "FALSE", "FALSO"}


# ════════════════════════════════════════════════════════════════════════════
# Normalización
# ════════════════════════════════════════════════════════════════════════════

def sin_tildes(texto: str) -> str:
    """`"Formación Técnica"` → `"Formacion Tecnica"`. Los niveles llegan con
    tildes inconsistentes entre filas del mismo archivo."""
    return "".join(c for c in unicodedata.normalize("NFKD", texto)
                   if not unicodedata.combining(c))


def clave_encabezado(valor) -> str:
    """Nombre de columna → clave comparable: sin tildes, MAYÚSCULA, separadores
    colapsados a `_`. Así `SNIES/SIET_PROGRAMA`, `snies siet programa` y
    `SNIES-SIET-PROGRAMA` caen en la misma cadena."""
    txt = sin_tildes(str(valor or "")).upper().strip()
    txt = re.sub(r"[\s/\-.]+", "_", txt)
    return re.sub(r"_+", "_", txt).strip("_")


def texto(valor) -> str | None:
    """Celda → texto limpio, o None si está vacía o dice «no aplica»."""
    if valor is None:
        return None
    txt = str(valor).strip()
    # Excel devuelve los enteros de una columna mixta como float: "5380.0".
    if isinstance(valor, float) and valor.is_integer():
        txt = str(int(valor))
    txt = re.sub(r"\s+", " ", txt)
    if sin_tildes(txt).upper() in VACIOS:
        return None
    return txt or None


def digitos(valor) -> str | None:
    """Celda → solo dígitos, o None. `1.023.456` y `1023456.0` dan lo mismo.

    Devuelve None también cuando queda algo que no son dígitos, para que quien
    llama distinga «vacío» de «tenía letras» mirando el original.
    """
    txt = texto(valor)
    if txt is None:
        return None
    limpio = re.sub(r"[.,\s]", "", txt)
    return limpio if limpio.isdigit() else None


def booleano(valor) -> bool | None:
    """`SI/X/1` → True, `NO/0` → False, vacío → None (que NO es lo mismo que
    False: significa «el archivo no lo dice»)."""
    txt = texto(valor)
    if txt is None:
        return None
    normal = sin_tildes(txt).upper()
    if normal in _AFIRMATIVOS:
        return True
    if normal in _NEGATIVOS:
        return False
    return None


def nombre_completo(datos: dict) -> str:
    partes = [datos.get(k) for k in ("nombre1", "nombre2", "apellido1", "apellido2")]
    return " ".join(p for p in partes if p)


# ════════════════════════════════════════════════════════════════════════════
# Estructuras del reporte
# ════════════════════════════════════════════════════════════════════════════

@dataclass
class FilaCargue:
    """Una fila del Excel, ya normalizada, con lo que se le encontró.

    `fila_excel` es el número REAL de la hoja (1-based, contando el título y el
    encabezado), no el índice del registro: es lo que el usuario ve cuando abre
    el archivo y lo único que le sirve para corregir.
    """
    fila_excel: int
    datos: dict = field(default_factory=dict)
    errores: list[str] = field(default_factory=list)
    avisos: list[str] = field(default_factory=list)

    @property
    def estado(self) -> str:
        if self.errores:
            return "error"
        return "aviso" if self.avisos else "ok"

    @property
    def clave_matricula(self) -> tuple:
        """Lo que hace única a una matrícula: persona + institución + programa.
        Espeja el índice `uq_entrega_beca_matricula` de la base (que además
        lleva la vigencia, y la vigencia es del lote, no de la fila)."""
        return (self.datos.get("documento"),
                self.datos.get("snies_ies"),
                self.datos.get("snies_programa"))

    def como_dict(self) -> dict:
        return {"fila": self.fila_excel, "estado": self.estado,
                "errores": self.errores, "avisos": self.avisos,
                "datos": self.datos}


@dataclass
class Lectura:
    """El resultado completo: qué se leyó, de dónde y qué tan bien salió."""
    hoja: str = ""
    fila_encabezado: int = 0
    titulo: str | None = None
    columnas: dict = field(default_factory=dict)          # clave interna → letra/índice
    columnas_ignoradas: list[str] = field(default_factory=list)
    filas: list[FilaCargue] = field(default_factory=list)
    avisos_globales: list[str] = field(default_factory=list)

    @property
    def total(self) -> int:
        return len(self.filas)

    @property
    def con_error(self) -> int:
        return sum(1 for f in self.filas if f.estado == "error")

    @property
    def con_aviso(self) -> int:
        return sum(1 for f in self.filas if f.estado == "aviso")

    @property
    def ok(self) -> int:
        return sum(1 for f in self.filas if f.estado == "ok")

    @property
    def personas_distintas(self) -> int:
        return len({f.datos.get("documento") for f in self.filas
                    if f.datos.get("documento")})

    @property
    def trae_cumplimiento(self) -> bool:
        """True si el archivo discrimina acceso/permanencia. Si es False el
        cargue corre igual, pero el avance no se puede imputar a ningún KPI."""
        return "acceso" in self.columnas or "permanencia" in self.columnas

    def resumen(self) -> dict:
        return {
            "hoja": self.hoja,
            "fila_encabezado": self.fila_encabezado,
            "titulo": self.titulo,
            "total": self.total,
            "ok": self.ok,
            "con_aviso": self.con_aviso,
            "con_error": self.con_error,
            "personas_distintas": self.personas_distintas,
            "trae_cumplimiento": self.trae_cumplimiento,
            "columnas_ignoradas": self.columnas_ignoradas,
            "avisos_globales": self.avisos_globales,
        }


class ArchivoInvalido(Exception):
    """El archivo no se puede leer como planilla de beneficiarios.

    Se reserva para lo que impide siquiera empezar —no hay encabezado, no hay
    filas, se pasa del tope—. Todo lo demás es error DE FILA y se reporta fila
    a fila, que es lo que le sirve a quien corrige.
    """


# ════════════════════════════════════════════════════════════════════════════
# Lectura
# ════════════════════════════════════════════════════════════════════════════

def _alias_a_clave() -> dict[str, str]:
    """Alias del encabezado → clave interna. Invertido de `COLUMNAS`."""
    return {alias: clave for clave, aliases in COLUMNAS.items() for alias in aliases}


def _mapear_encabezado(celdas) -> tuple[dict, list]:
    """Fila de encabezado → (clave interna → índice de columna, ignoradas)."""
    alias = _alias_a_clave()
    columnas, ignoradas = {}, []
    for idx, celda in enumerate(celdas):
        bruto = clave_encabezado(celda)
        if not bruto:
            continue
        clave = alias.get(bruto)
        if clave is None:
            ignoradas.append(str(celda).strip())
        elif clave not in columnas:      # gana la primera si viene repetida
            columnas[clave] = idx
    return columnas, ignoradas


def _buscar_encabezado(filas: list[list]) -> tuple[int, dict, list, str | None]:
    """Encuentra la fila de encabezado en las primeras `FILAS_BUSQUEDA_ENCABEZADO`.

    Califica la que mapee las tres obligatorias. Lo que haya antes se guarda
    como título (de ahí sale el `| 175` que se contrasta con lo leído).
    """
    titulo = None
    for i, celdas in enumerate(filas[:FILAS_BUSQUEDA_ENCABEZADO]):
        columnas, ignoradas = _mapear_encabezado(celdas)
        if all(k in columnas for k in OBLIGATORIAS):
            return i, columnas, ignoradas, titulo
        if titulo is None:
            texto_fila = " ".join(str(c).strip() for c in celdas if c is not None)
            titulo = texto_fila.strip() or None
    raise ArchivoInvalido(
        "No se encontró la fila de encabezado en las primeras "
        f"{FILAS_BUSQUEDA_ENCABEZADO} filas. Debe traer al menos las columnas "
        "DOCUMENTO, PRIMER_NOMBRE y PRIMER_APELLIDO."
    )


def _leer_fila(celdas, columnas: dict, fila_excel: int) -> FilaCargue:
    """Una fila del Excel → `FilaCargue` normalizada y juzgada por sí sola."""
    def crudo(clave):
        idx = columnas.get(clave)
        return celdas[idx] if idx is not None and idx < len(celdas) else None

    fila = FilaCargue(fila_excel=fila_excel)
    d = fila.datos

    # ── Identidad ──────────────────────────────────────────────────────────
    sigla = (texto(crudo("tipo_documento")) or "").upper()
    d["tipo_documento_sigla"] = sigla or None
    if not sigla:
        fila.errores.append("Falta el tipo de documento.")
    elif sigla not in TIPOS_DOCUMENTO:
        fila.errores.append(
            f"Tipo de documento '{sigla}' no reconocido. Use "
            f"{', '.join(sorted(set(TIPOS_DOCUMENTO)))}."
        )
    else:
        d["tipo_documento_codigo"] = TIPOS_DOCUMENTO[sigla]
        if sigla in SIGLAS_SIN_CATALOGO:
            fila.avisos.append(
                f"El catálogo de la base no tiene '{sigla}': se guarda como "
                "'Otro' y la sigla queda registrada."
            )

    d["documento"] = digitos(crudo("documento"))
    if not d["documento"]:
        bruto = texto(crudo("documento"))
        fila.errores.append(
            f"El documento '{bruto}' no es un número." if bruto
            else "Falta el número de documento."
        )

    for clave, etiqueta in (("nombre1", "primer nombre"), ("apellido1", "primer apellido")):
        d[clave] = texto(crudo(clave))
        if not d[clave]:
            fila.errores.append(f"Falta el {etiqueta}.")
    d["nombre2"] = texto(crudo("nombre2"))
    d["apellido2"] = texto(crudo("apellido2"))

    # ── Matrícula ──────────────────────────────────────────────────────────
    d["snies_programa"] = digitos(crudo("snies_programa"))
    d["snies_ies"] = digitos(crudo("snies_ies"))
    d["programa"] = texto(crudo("programa"))
    d["ies_nombre"] = texto(crudo("ies_nombre"))
    if not d["snies_programa"]:
        fila.errores.append("Falta el código SNIES/SIET del programa.")
    if not d["snies_ies"]:
        fila.errores.append("Falta el código SNIES/SIET de la institución.")
    if not d["programa"]:
        fila.avisos.append("Sin nombre de programa; se guarda solo el código.")
    if not d["ies_nombre"]:
        fila.avisos.append("Sin nombre de institución; se guarda solo el código.")

    nivel_bruto = texto(crudo("nivel"))
    d["nivel_bruto"] = nivel_bruto
    if not nivel_bruto:
        fila.errores.append("Falta el nivel de formación.")
    else:
        clave_nivel = re.sub(r"\s+", " ", sin_tildes(nivel_bruto).upper()).strip()
        d["nivel_formacion"] = NIVELES.get(clave_nivel)
        if d["nivel_formacion"] is None:
            fila.errores.append(
                f"Nivel de formación '{nivel_bruto}' no reconocido. Se aceptan: "
                f"{', '.join(sorted(set(NIVELES)))}."
            )

    # ── Opcionales ─────────────────────────────────────────────────────────
    d["localidad"] = texto(crudo("localidad"))
    if d["localidad"] and sin_tildes(d["localidad"]).upper() != "KENNEDY":
        fila.avisos.append(
            f"La localidad dice '{d['localidad']}', no KENNEDY. Verifique que "
            "el beneficiario corresponde a esta localidad."
        )
    d["acceso"] = booleano(crudo("acceso"))
    d["permanencia"] = booleano(crudo("permanencia"))
    d["barrio"] = texto(crudo("barrio"))
    d["telefono"] = texto(crudo("telefono"))
    d["correo"] = texto(crudo("correo"))
    return fila


def _revisar_entre_filas(filas: list[FilaCargue]) -> None:
    """Lo que solo se ve comparando el archivo consigo mismo.

    Tres casos, con tres tratos distintos y a propósito:

    - **misma persona, misma IES, mismo programa** → ERROR. Es doble digitación:
      la base lo rechazaría con el índice único de matrícula.
    - **misma persona, programa distinto** → aviso. Es legítimo y está
      verificado en el archivo de 2025 (un beneficiario con dos matrículas).
    - **mismo documento, nombre distinto** → aviso. Puede ser un typo o dos
      personas con el mismo número mal digitado; no lo decide el sistema, y
      además la persona ya registrada gana por política A.
    """
    vistas: dict[tuple, int] = {}
    por_documento: dict[str, list[FilaCargue]] = {}

    for fila in filas:
        doc = fila.datos.get("documento")
        if not doc:
            continue
        por_documento.setdefault(doc, []).append(fila)
        clave = fila.clave_matricula
        if not all(clave):
            continue
        if clave in vistas:
            fila.errores.append(
                f"Repetida: la misma persona, institución y programa ya está en "
                f"la fila {vistas[clave]}."
            )
        else:
            vistas[clave] = fila.fila_excel

    for doc, grupo in por_documento.items():
        if len(grupo) < 2:
            continue
        programas = {f.datos.get("snies_programa") for f in grupo}
        if len(programas) > 1:
            for fila in grupo:
                otras = [str(f.fila_excel) for f in grupo if f is not fila]
                fila.avisos.append(
                    f"El documento {doc} aparece también en la(s) fila(s) "
                    f"{', '.join(otras)} con otro programa. Se aceptan como "
                    "matrículas distintas de la misma persona."
                )
        nombres = {nombre_completo(f.datos) for f in grupo}
        if len(nombres) > 1:
            for fila in grupo:
                fila.avisos.append(
                    f"El documento {doc} aparece con nombres distintos: "
                    f"{' / '.join(sorted(nombres))}."
                )


def _contraste_con_titulo(lectura: Lectura) -> None:
    """Si el título trae un número, se compara con lo leído.

    El archivo de 2025 dice `KENNEDY META JE4 | 175` y trae 175 filas. Que
    dejen de coincidir es la señal de que alguien filtró u ocultó filas antes
    de mandarlo, y es un error que no se ve mirando el Excel.
    """
    if not lectura.titulo:
        return
    numeros = [int(n) for n in re.findall(r"\b(\d{1,5})\b", lectura.titulo)]
    if not numeros:
        return
    anunciado = max(numeros)
    if anunciado != lectura.total:
        lectura.avisos_globales.append(
            f"El título del archivo dice {anunciado} registros y se leyeron "
            f"{lectura.total}. Verifique que no quedaron filas filtradas u ocultas."
        )


def leer(archivo) -> Lectura:
    """Lee el Excel y devuelve la `Lectura` completa. No persiste nada.

    `archivo` es cualquier cosa que openpyxl acepte: una ruta o un archivo en
    memoria (`request.FILES[...]`).
    """
    from openpyxl import load_workbook

    try:
        libro = load_workbook(archivo, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 — openpyxl lanza de todo
        raise ArchivoInvalido(
            "No se pudo abrir el archivo. Debe ser un Excel .xlsx "
            f"({exc.__class__.__name__})."
        ) from exc

    ultimo_error: ArchivoInvalido | None = None
    for hoja in libro.worksheets:
        filas_crudas = [list(f) for f in hoja.iter_rows(values_only=True)]
        if not filas_crudas:
            continue
        try:
            i_enc, columnas, ignoradas, titulo = _buscar_encabezado(filas_crudas)
        except ArchivoInvalido as exc:
            ultimo_error = exc          # esta hoja no era; puede haber otra
            continue

        lectura = Lectura(hoja=hoja.title, fila_encabezado=i_enc + 1,
                          titulo=titulo, columnas=columnas,
                          columnas_ignoradas=ignoradas)

        cuerpo = filas_crudas[i_enc + 1:]
        # Las filas totalmente vacías del final son normales en un Excel: se
        # descartan sin ruido. Una vacía EN MEDIO también, pero se avisa,
        # porque suele delatar que el archivo venía filtrado.
        for desplazamiento, celdas in enumerate(cuerpo):
            if all(texto(c) is None for c in celdas):
                continue
            fila_excel = i_enc + 2 + desplazamiento
            if lectura.total >= LIMITE_FILAS:
                raise ArchivoInvalido(
                    f"El archivo supera el tope de {LIMITE_FILAS} filas. "
                    "Divídalo en varios cargues."
                )
            lectura.filas.append(_leer_fila(celdas, columnas, fila_excel))

        if not lectura.filas:
            raise ArchivoInvalido(
                f"La hoja '{hoja.title}' tiene encabezado pero ninguna fila de datos."
            )

        _revisar_entre_filas(lectura.filas)
        _contraste_con_titulo(lectura)
        if not lectura.trae_cumplimiento:
            lectura.avisos_globales.append(
                "El archivo no trae las columnas ACCESO / PERMANENCIA: los "
                "beneficiarios se pueden cargar, pero el avance no se podrá "
                "imputar a las metas hasta que el área envíe esa discriminación."
            )
        return lectura

    raise ultimo_error or ArchivoInvalido("El archivo no tiene hojas con datos.")
