"""Vistas de organizador del módulo Jóvenes a la E (login + permisos).

J1 — Listado paginado con filtros (estado / evento / búsqueda por
     documento o nombre).
J1 — Detalle de una entrega con todos los campos + lista de elementos.
J1 — Validar / rechazar con observación.
J2 — Al **validar**, se sincronizan los avances en
     `presu_avance_ind_periodo`: +1 al KPI vinculado a la actividad_plan
     del evento por cada cumplimiento marcado (acceso 23771 / permanencia 23772).
     Al **rechazar** una entrega que ya estaba validada, los avances se
     revierten (delete de las filas con origen='EVENTO' que apunten a
     este evento+indicador+entrega).
J5 — Insights (Chart.js) + descarga Excel con Matriz 1 (presupuestal)
     y Matriz 2 (ejecución contractual). Patrón Banco.
"""
import io
import logging
from datetime import date

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import redirect

from apps.login.decorators import modulo_required, jwt_or_session_required
from apps.jovenes_a_la_e.models import EntregaBeca


logger = logging.getLogger(__name__)


# Mapa cumplimiento → meta_codigo (planilla externa)
META_POR_CUMPLIMIENTO = {
    "cumplimiento_acceso": "23771",
    "cumplimiento_permanencia": "23772",
}


# ─────────────────────────────────────────────────────────────────
# Helpers J2 — sincronización con AvanceIndicador
# ─────────────────────────────────────────────────────────────────

def _indicadores_del_evento(evento) -> list:
    """Lista de Indicador (KPI) ligados a la actividad_plan del evento.

    Vacío si el evento no tiene actividad_plan o si esa actividad no
    tiene KPIs vinculados — en ese caso la validación no sincroniza
    nada (no es error, solo no hay a qué sumar).
    """
    if not evento or not evento.actividad_plan_id:
        return []
    from apps.presupuesto.models import ActividadIndicador
    return list(
        ActividadIndicador.objects
        .filter(actividad_plan_id=evento.actividad_plan_id)
        .select_related("indicador")
    )


def _sincronizar_avance(entrega: EntregaBeca, *, accion: str):
    """J2: al validar suma avance, al rechazar (después de validada) revierte.

    Estrategia simple: una fila en `presu_avance_ind_periodo` por cada
    par (entrega, indicador, cumplimiento). Se identifica por
    `origen='EVENTO'` + `evento_id` + `indicador_id`. Para revertir se
    borra y listo (no se hace rollback de magnitud porque la magnitud
    aquí siempre es 1 por cumplimiento marcado).
    """
    from apps.presupuesto.models import AvanceIndicador

    relaciones = _indicadores_del_evento(entrega.evento)
    if not relaciones:
        return 0

    cumplimientos = []
    if entrega.cumplimiento_acceso:       cumplimientos.append("23771")
    if entrega.cumplimiento_permanencia:  cumplimientos.append("23772")
    if not cumplimientos:
        return 0

    fecha_aporte = date.today()
    periodo = fecha_aporte.strftime("%Y-%m")
    n = 0

    if accion == "validar":
        for rel in relaciones:
            ind = rel.indicador
            # Hacemos idempotente: solo creamos si no hay ya un avance
            # de esta entrega+indicador (la entrega podría re-validarse).
            ya = AvanceIndicador.objects.filter(
                indicador_id=ind.id,
                evento_id=entrega.evento_id,
                origen="EVENTO",
                observaciones__contains=f"entrega_beca={entrega.id}",
            ).exists()
            if ya:
                continue
            AvanceIndicador.objects.create(
                indicador_id=ind.id,
                evento_id=entrega.evento_id,
                magnitud_aportada=len(cumplimientos),  # 1 o 2 según cumplimientos
                fecha_aporte=fecha_aporte,
                periodo=periodo,
                origen="EVENTO",
                observaciones=f"entrega_beca={entrega.id}; metas={','.join(cumplimientos)}",
            )
            n += 1

    elif accion == "revertir":
        for rel in relaciones:
            borrados = AvanceIndicador.objects.filter(
                indicador_id=rel.indicador.id,
                evento_id=entrega.evento_id,
                origen="EVENTO",
                observaciones__contains=f"entrega_beca={entrega.id}",
            ).delete()
            n += borrados[0] if borrados else 0

    return n


# ─────────────────────────────────────────────────────────────────
# Vistas organizador
# ─────────────────────────────────────────────────────────────────

@login_required
@modulo_required("jovenes_a_la_e")
def entregas_list(request):
    """Migrado a Angular: listado de entregas de becas."""
    return redirect("/app/jovenes")


# ─────────────────────────────────────────────────────────────────
# J5 — Insights + Excel
# ─────────────────────────────────────────────────────────────────

# Meta del proyecto 2805 "Kennedy Germinando Futuros" — convenio 773-2025:
#   23771 acceso       → 700 estudiantes posmedia
#   23772 permanencia  → 700 estudiantes posmedia
# Total combinado = 1400 cumplimientos (un estudiante puede cumplir ambos).
META_ACCESO = 700
META_PERMANENCIA = 700
META_TOTAL = META_ACCESO + META_PERMANENCIA


@login_required
@modulo_required("jovenes_a_la_e")
def entregas_insights(request):
    """Migrado a Angular: insights de Jóvenes a la E."""
    return redirect("/app/jovenes/insights")


@jwt_or_session_required
@modulo_required("jovenes_a_la_e")
def entregas_exportar_excel(request):
    """Descarga Excel xlsx con 4 hojas:

    1. Resumen — KPIs trascendentales.
    2. Entregas detalle — una fila por entrega con todos los campos.
    3. Matriz 1 (Presupuestal) — Proyecto → Meta → KPI → Avance.
    4. Matriz 2 (Ejecución contractual) — CDP → Contrato → Vinculación
       → ActividadPlan → Eventos → Entregas.

    Mantenemos la estructura estándar Alex (memoria
    `feedback_matrices_estandar`): toda la cadena debe quedar ligada.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()

    # ── Estilos compartidos ───────────────────────────────────
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _write_header(ws, columns):
        for col, name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col, value=name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        ws.row_dimensions[1].height = 28

    # ── Hoja 1: Resumen ───────────────────────────────────────
    ws = wb.active
    ws.title = "Resumen"
    qs = EntregaBeca.objects.all()
    total = qs.count()
    validadas = qs.filter(estado="validada").count()
    rechazadas = qs.filter(estado="rechazada").count()
    enviadas = qs.filter(estado="enviada").count()
    cumple_acceso = qs.filter(estado="validada", cumplimiento_acceso=True).count()
    cumple_permanencia = qs.filter(estado="validada", cumplimiento_permanencia=True).count()
    con_firma = qs.filter(firma_mongo_id__isnull=False).count()

    resumen_filas = [
        ("Indicador", "Valor"),
        ("Total entregas registradas", total),
        ("Enviadas (sin validar)", enviadas),
        ("Validadas", validadas),
        ("Rechazadas", rechazadas),
        ("% validación cerrada", f"{round(100*(validadas+rechazadas)/total,1) if total else 0}%"),
        ("", ""),
        ("Meta acceso (23771)", META_ACCESO),
        ("Cumplimiento acceso (validadas)", cumple_acceso),
        ("% avance acceso", f"{round(100*cumple_acceso/META_ACCESO,1) if META_ACCESO else 0}%"),
        ("", ""),
        ("Meta permanencia (23772)", META_PERMANENCIA),
        ("Cumplimiento permanencia (validadas)", cumple_permanencia),
        ("% avance permanencia", f"{round(100*cumple_permanencia/META_PERMANENCIA,1) if META_PERMANENCIA else 0}%"),
        ("", ""),
        ("Calidad — con firma cifrada", con_firma),
        ("% con firma", f"{round(100*con_firma/total,1) if total else 0}%"),
    ]
    _write_header(ws, list(resumen_filas[0]))
    for i, (k, v) in enumerate(resumen_filas[1:], start=2):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)
    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 18

    # ── Hoja 2: Entregas detalle ──────────────────────────────
    ws2 = wb.create_sheet("Entregas detalle")
    cols = [
        "ID", "Estado", "Fecha captura", "Evento", "Convenio", "Proyecto",
        "Tipo doc", "Nº documento", "Nombres", "Apellidos",
        "Teléfono", "Correo", "Dirección", "Barrio", "UPL",
        "Cumple acceso", "Cumple permanencia", "Nivel formación",
        "Institución", "Programa académico", "Período",
        "Tiene firma (Mongo)", "Observaciones",
    ]
    _write_header(ws2, cols)
    row = 2
    for e in qs.select_related("evento"):
        nombres = " ".join(filter(None, [e.nombre1, e.nombre2]))
        apellidos = " ".join(filter(None, [e.apellido1, e.apellido2]))
        upl_nom = ""
        if e.upl_codigo:
            from apps.banco_iniciativas.models.catalogos import Upl
            u = Upl.objects.filter(codigo=e.upl_codigo).first()
            upl_nom = u.nombre if u else f"UPL {e.upl_codigo}"
        ws2.append([
            e.id, e.estado, e.created_at.strftime("%Y-%m-%d %H:%M") if e.created_at else "",
            e.evento.nombre if e.evento_id else "",
            e.convenio_codigo, e.proyecto_codigo,
            e.tipo_doc_codigo or "", e.numero_documento,
            nombres, apellidos,
            e.telefono or "", e.correo or "",
            e.direccion or "", e.barrio_codigo or "", upl_nom,
            "Sí" if e.cumplimiento_acceso else "No",
            "Sí" if e.cumplimiento_permanencia else "No",
            dict(EntregaBeca.NIVEL_CHOICES).get(e.nivel_formacion, e.nivel_formacion or ""),
            e.institucion or "", e.programa_academico or "", e.periodo_academico or "",
            "Sí" if e.firma_mongo_id else "No",
            (e.observaciones or "")[:200],
        ])
        row += 1
    for col_idx in range(1, len(cols) + 1):
        ws2.column_dimensions[ws2.cell(row=1, column=col_idx).column_letter].width = 18

    # ── Hoja 3: Matriz 1 — Presupuestal ───────────────────────
    # Proyecto → Meta → KPI (Indicador) → Avance acumulado
    from django.db.models import Sum
    from apps.presupuesto.models import ActividadIndicador, AvanceIndicador
    from apps.login.models.evento import Evento
    ws3 = wb.create_sheet("Matriz 1 Presupuestal")
    _write_header(ws3, [
        "Proyecto código", "Proyecto nombre", "Meta código", "Meta descripción",
        "KPI / Indicador", "Unidad", "Meta magnitud",
        "Avance acumulado", "% avance", "Origen avances",
    ])
    # Eventos del módulo Jóvenes a la E: JOVENES_BECA + (futuro) JOVENES_DOTACION_SEDE
    eventos_jovenes = Evento.objects.filter(
        tipo_evento__codigo__in=["JOVENES_BECA", "JOVENES_DOTACION_SEDE"]
    ).select_related("actividad_plan")
    actividad_ids = {e.actividad_plan_id for e in eventos_jovenes if e.actividad_plan_id}

    # Indicadores vinculados a esas actividades
    rels = (
        ActividadIndicador.objects
        .filter(actividad_plan_id__in=actividad_ids)
        .select_related("indicador", "indicador__meta_proyecto",
                        "indicador__meta_proyecto__meta",
                        "indicador__meta_proyecto__proyecto")
    )
    matriz1_rows = []
    for rel in rels:
        ind = rel.indicador
        mp = ind.meta_proyecto if ind else None
        meta = mp.meta if mp else None
        proy = mp.proyecto if mp else None
        # Suma acumulada de magnitudes (un solo query agregado)
        suma = AvanceIndicador.objects.filter(
            indicador=ind, activo=True
        ).aggregate(s=Sum("magnitud_aportada"))["s"] or 0
        pct = round(100 * float(suma) / float(ind.meta_magnitud), 1) if (ind and ind.meta_magnitud) else None
        matriz1_rows.append([
            (proy.codigo if proy else "") or "",
            (proy.nombre if proy else "") or "",
            (meta.codigo if meta else "") or "",
            (meta.nombre if meta else "") or "",  # MetaBD tiene `nombre`, no `descripcion`
            (ind.nombre if ind else "") or "",
            (ind.unidad_medida if ind else "") or "",
            float(ind.meta_magnitud) if ind and ind.meta_magnitud else 0,
            float(suma),
            pct if pct is not None else "",
            "EVENTO",
        ])
    # Dedupe + ordenar
    seen = set()
    for row_vals in matriz1_rows:
        key = tuple(row_vals[:5])
        if key in seen:
            continue
        seen.add(key)
        ws3.append(row_vals)
    for col_idx in range(1, 11):
        ws3.column_dimensions[ws3.cell(row=1, column=col_idx).column_letter].width = 22

    # ── Hoja 4: Matriz 2 — Ejecución contractual ──────────────
    # CDP → Contrato → ContratoActividadPlan → ActividadPlan → # eventos / # entregas
    from apps.presupuesto.models.sql import ContratoActividadPlan
    ws4 = wb.create_sheet("Matriz 2 Ejec. contractual")
    _write_header(ws4, [
        "CDP número", "Contrato número", "Contrato vigencia", "Contrato valor",
        "Actividad plan", "Monto vinculación",
        "# eventos Jóvenes", "# entregas registradas", "# validadas",
    ])
    vincs = (
        ContratoActividadPlan.objects
        .filter(actividad_plan_id__in=actividad_ids, activo=True)
        .select_related("contrato", "contrato__cdp", "actividad_plan")
    )
    for v in vincs:
        c = v.contrato
        ev_ids = [e.id for e in eventos_jovenes if e.actividad_plan_id == v.actividad_plan_id]
        n_ev = len(ev_ids)
        n_ent = EntregaBeca.objects.filter(evento_id__in=ev_ids).count()
        n_val = EntregaBeca.objects.filter(evento_id__in=ev_ids, estado="validada").count()
        ws4.append([
            (c.cdp.cdp_numero if c and c.cdp_id else "") or "",
            (c.contrato_numero if c else "") or "",
            (c.contrato_vigencia if c else "") or "",
            float(c.valor) if c and c.valor else 0,
            (v.actividad_plan.descripcion if v.actividad_plan_id else "") or f"AP {v.actividad_plan_id}",
            float(v.monto) if v.monto else 0,
            n_ev, n_ent, n_val,
        ])
    for col_idx in range(1, 10):
        ws4.column_dimensions[ws4.cell(row=1, column=col_idx).column_letter].width = 22

    # ── Devolver el archivo ───────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    nombre = f"jovenes_a_la_e_{date.today().isoformat()}.xlsx"
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{nombre}"'
    return response


@login_required
@modulo_required("jovenes_a_la_e")
def entrega_detalle(request, pk: int):
    """Migrado a Angular: detalle de entrega."""
    return redirect(f"/app/jovenes/{pk}")
