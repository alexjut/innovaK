"""Smoke tests del módulo Jóvenes a la E (PR-1, scope sólo Becas).

Cobertura mínima:
  - Imports OK (modelos + views + urls).
  - URLs registradas resuelven.
  - Módulo `jovenes_a_la_e` está en el catálogo `modulo` (post seed).
  - Tipo de evento `JOVENES_BECA` existe (post DDL).

Los tests usan el superuser real de BD vía Test Client. Si el DDL aún
no se ha aplicado, los tests dependientes se skippean — no fallan.
"""
import unittest

from django.conf import settings
from django.contrib.auth import get_user_model
from django.test import Client
from django.urls import reverse


HOST = settings.ALLOWED_HOSTS[0] if settings.ALLOWED_HOSTS else "localhost"


class JovenesALaESmokeTests(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        User = get_user_model()
        cls.user = User.objects.filter(is_superuser=True).first()
        if cls.user is None:
            raise unittest.SkipTest("No hay superuser en la BD")
        cls.client_auth = Client(HTTP_HOST=HOST)
        cls.client_auth.force_login(cls.user)
        cls.client_anon = Client(HTTP_HOST=HOST)

    # ── Imports ────────────────────────────────────────────────

    def test_modelos_importan(self):
        from apps.jovenes_a_la_e.models import (
            ElementoDotacion, EntregaBeca, EntregaBecaElemento,
        )
        self.assertEqual(ElementoDotacion._meta.db_table, "elemento_dotacion")
        self.assertEqual(EntregaBeca._meta.db_table, "entrega_beca")
        self.assertEqual(EntregaBecaElemento._meta.db_table, "entrega_beca_elemento")

    def test_modelos_no_managed(self):
        from apps.jovenes_a_la_e.models import (
            ElementoDotacion, EntregaBeca, EntregaBecaElemento,
        )
        for m in (ElementoDotacion, EntregaBeca, EntregaBecaElemento):
            self.assertFalse(m._meta.managed, f"{m.__name__} debe ser managed=False")

    # ── URLs ───────────────────────────────────────────────────

    def test_urls_resuelven(self):
        self.assertEqual(
            reverse("jovenes_a_la_e:form_publico_beca", args=[1]),
            "/jovenes-a-la-e/1/beca/",
        )
        self.assertEqual(
            reverse("jovenes_a_la_e:entrega_exitosa", args=[1]),
            "/jovenes-a-la-e/exitoso/1/",
        )
        self.assertEqual(
            reverse("jovenes_a_la_e:entregas_list"),
            "/jovenes-a-la-e/entregas/",
        )
        self.assertEqual(
            reverse("jovenes_a_la_e:entrega_detalle", args=[1]),
            "/jovenes-a-la-e/entregas/1/",
        )

    def test_form_publico_responde_para_evento_valido(self):
        """Migrado a Angular: la vista vieja redirige al form público SPA."""
        from apps.login.models import Evento
        evento = (
            Evento.objects
            .filter(activo=True, tipo_evento__codigo="JOVENES_BECA")
            .order_by("-id").first()
        )
        if evento is None:
            self.skipTest("No hay evento JOVENES_BECA activo en BD.")
            return
        r = self.client_anon.get(f"/jovenes-a-la-e/{evento.id}/beca/")
        # El form público migró a Angular (/app/p/jovenes/<id>): la URL Django
        # redirige (302) a la SPA, sin pasar por login (sigue siendo público).
        self.assertEqual(r.status_code, 302)
        self.assertIn(f"/app/p/jovenes/{evento.id}", r["Location"])
        self.assertNotIn("/login", r["Location"])

    def test_form_publico_evento_inexistente_404(self):
        r = self.client_anon.get("/jovenes-a-la-e/99999999/beca/")
        self.assertEqual(r.status_code, 404)

    def test_form_publico_evento_tipo_distinto_404(self):
        """Un evento de tipo distinto a JOVENES_BECA no debe servir este form."""
        from apps.login.models import Evento
        otro = (
            Evento.objects
            .exclude(tipo_evento__codigo="JOVENES_BECA")
            .filter(activo=True)
            .first()
        )
        if otro is None:
            self.skipTest("No hay otros eventos activos para probar el guard.")
            return
        r = self.client_anon.get(f"/jovenes-a-la-e/{otro.id}/beca/")
        self.assertEqual(r.status_code, 404)

    # ── Catálogo de módulos ───────────────────────────────────

    def test_modulo_jovenes_en_catalogo(self):
        try:
            from apps.login.models.permisos import Modulo, RolModulo
        except Exception:
            self.skipTest("Modelo Modulo no disponible.")
            return
        m = Modulo.objects.filter(codigo="jovenes_a_la_e").first()
        if m is None:
            self.skipTest("seed_modulos aún no se ha corrido tras agregar jovenes_a_la_e.")
            return
        self.assertTrue(m.activo)
        nombres = set(
            RolModulo.objects.filter(modulo=m)
            .values_list("group__name", flat=True)
        )
        self.assertIn("Admin", nombres)
        self.assertIn("Lider", nombres)

    # ── DDL aplicado ───────────────────────────────────────────

    def test_tipo_evento_jovenes_beca_existe(self):
        from apps.login.models.evento import TipoEvento
        if not TipoEvento.objects.filter(codigo="JOVENES_BECA").exists():
            self.skipTest("DDL 001_jovenes_setup.sql aún no aplicado.")
            return
        t = TipoEvento.objects.get(codigo="JOVENES_BECA")
        self.assertTrue(t.activo)
        self.assertTrue(t.permite_inscripcion)
        self.assertTrue(t.permite_qr)

    # ── J5: Insights + Excel ──────────────────────────────────

    def test_insights_responde_200(self):
        r = self.client_auth.get("/jovenes-a-la-e/entregas/insights/")
        self.assertEqual(r.status_code, 200)
        html = r.content.decode("utf-8", errors="ignore")
        self.assertIn("Insights", html)
        # Metas 23771 y 23772 deben estar referenciadas en el dashboard.
        self.assertIn("23771", html)
        self.assertIn("23772", html)
        # Chart.js debe cargarse en la página.
        self.assertIn("chart.umd", html.lower() if "chart.umd" in html else "chart.js")

    def test_insights_requiere_autenticacion(self):
        r = self.client_anon.get("/jovenes-a-la-e/entregas/insights/")
        # @login_required redirige a /login/?next=... (302)
        self.assertIn(r.status_code, (302, 403))

    def test_exportar_excel_devuelve_xlsx(self):
        r = self.client_auth.get("/jovenes-a-la-e/entregas/exportar/")
        self.assertEqual(r.status_code, 200)
        self.assertIn("spreadsheetml.sheet", r.get("Content-Type", ""))
        self.assertIn("jovenes_a_la_e_", r.get("Content-Disposition", ""))
        # Es un xlsx válido
        from openpyxl import load_workbook
        import io
        wb = load_workbook(io.BytesIO(r.content))
        for hoja in ("Resumen", "Entregas detalle",
                     "Matriz 1 Presupuestal", "Matriz 2 Ejec. contractual"):
            self.assertIn(hoja, wb.sheetnames, f"Falta la hoja {hoja}")

    def test_exportar_excel_requiere_autenticacion(self):
        r = self.client_anon.get("/jovenes-a-la-e/entregas/exportar/")
        self.assertIn(r.status_code, (302, 403))
