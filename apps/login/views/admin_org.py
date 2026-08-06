# apps/login/views/admin_org.py
"""
CRUD para entidades organizativas:
  - PR-F: Dependencia, Subgrupo, Funcionario.
  - PR-H2: Organización, Proveedor, Beneficiario.
"""
from django.contrib.auth.decorators import login_required
from django.shortcuts import redirect

# Tamaño de página por defecto para listados largos (P1).
PAGE_SIZE = 25

from apps.login.decorators import modulo_required, jwt_or_session_required
from apps.login.models.funcionario import (
    Dependencia, Subgrupo, Funcionario, TipoFuncionario, Cargo
)
from apps.login.models.persona import Persona
from apps.login.models.contratos import (
    Organizacion, Proveedor, Beneficiario,
)


# ──────────────────────────────────────────────────────────────
# Forms
# ──────────────────────────────────────────────────────────────

# ──────────────────────────────────────────────────────────────
# CRUD Dependencia
# ──────────────────────────────────────────────────────────────

@login_required
@modulo_required('org_admin')
def dependencias_list(request):
    """Migrado a Angular: gestión de organización."""
    return redirect("/app/admin/org")


# ──────────────────────────────────────────────────────────────
# CRUD Subgrupo
# ──────────────────────────────────────────────────────────────

@login_required
@modulo_required('org_admin')
def subgrupos_list(request):
    """Migrado a Angular: gestión de organización."""
    return redirect("/app/admin/org")


# ──────────────────────────────────────────────────────────────
# CRUD Funcionario
# ──────────────────────────────────────────────────────────────

@login_required
@modulo_required('org_admin')
def funcionarios_list(request):
    """Migrado a Angular: gestión de organización."""
    return redirect("/app/admin/org")


@login_required
@modulo_required('org_admin')
def funcionario_nuevo(request):
    """Migrado a Angular: alta de funcionario (form inline)."""
    return redirect("/app/admin/org")


# ══════════════════════════════════════════════════════════════
#   PR-H2: Organización · Proveedor · Beneficiario
# ══════════════════════════════════════════════════════════════

# ──────────────────────────────────────────────────────────────
# Forms
# ──────────────────────────────────────────────────────────────

# ──────────────────────────────────────────────────────────────
# CRUD Organización
# ──────────────────────────────────────────────────────────────

@login_required
@modulo_required('org_admin')
def organizaciones_list(request):
    """Migrado a Angular: gestión de organización."""
    return redirect("/app/admin/org")


# ──────────────────────────────────────────────────────────────
# CRUD Proveedor
# ──────────────────────────────────────────────────────────────

@login_required
@modulo_required('org_admin')
def proveedores_list(request):
    """Migrado a Angular: gestión de organización."""
    return redirect("/app/admin/org")


# ──────────────────────────────────────────────────────────────
# CRUD Beneficiario
# ──────────────────────────────────────────────────────────────

@login_required
@modulo_required('org_admin')
def beneficiarios_list(request):
    """Migrado a Angular: gestión de organización."""
    return redirect("/app/admin/org")


@jwt_or_session_required
@modulo_required('org_admin')
def beneficiarios_exportar_excel(request):
    """Descarga XLSX nativo de beneficiarios. Respeta filtro `?tipo=X`.

    Diseñado para entregar a coordinadores y comunicaciones: archivo
    Excel listo (no CSV con extensión), con header congelado, ancho de
    columnas ajustado, banding alternado.
    """
    from io import BytesIO
    from datetime import datetime
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    qs = (
        Beneficiario.objects
        .select_related("persona", "proveedor", "organizacion", "tipo_documento")
        .filter(activo=True)
    )
    tipo_filter = (request.GET.get("tipo") or "").strip().upper()
    if tipo_filter in {"PERSONA", "ORGANIZACION", "PROVEEDOR"}:
        qs = qs.filter(tipo=tipo_filter)
    qs = qs.order_by("tipo", "nombre_legal")

    wb = Workbook()
    ws = wb.active
    ws.title = f"Beneficiarios{(' ' + tipo_filter.title()) if tipo_filter else ''}"

    headers = [
        "ID", "Tipo", "Tipo Documento", "Número Documento", "Nombre Legal",
        "Correo", "Teléfono", "Dirección",
        "Persona ID", "Persona Nombre",
        "Organización ID", "Organización Nombre", "Organización NIT",
        "Proveedor ID", "Proveedor Nombre",
        "Activo",
    ]
    # Estilos institucionales Alcaldía
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="D6001C")  # rojo Alcaldía
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="E5E7EB")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    band_fill = PatternFill("solid", fgColor="F9FAFB")

    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = cell_border
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    for r_idx, b in enumerate(qs.iterator(chunk_size=500), start=2):
        persona_nombre = ""
        if b.persona:
            parts = [b.persona.nombre1 or "", b.persona.nombre2 or "",
                     b.persona.apellido1 or "", b.persona.apellido2 or ""]
            persona_nombre = " ".join(p for p in parts if p).strip()
        ws.append([
            b.id,
            b.tipo or "",
            getattr(b.tipo_documento, "nombre", "") or "",
            b.numero_documento or "",
            b.nombre_legal or "",
            b.correo or "",
            b.telefono or "",
            b.direccion or "",
            b.persona_id or "",
            persona_nombre,
            b.organizacion_id or "",
            getattr(b.organizacion, "nombre", "") or "",
            getattr(b.organizacion, "nit", "") or "",
            b.proveedor_id or "",
            getattr(b.proveedor, "nombre", "") or "",
            "Sí" if b.activo else "No",
        ])
        # Banding alternado + bordes
        if r_idx % 2 == 0:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=r_idx, column=col_idx).fill = band_fill

    # Ancho de columnas inteligente
    anchos = [8, 14, 18, 18, 36, 28, 14, 36, 10, 30, 14, 30, 16, 12, 25, 8]
    for i, w in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Hoja resumen
    resumen = wb.create_sheet("Resumen")
    total = qs.count()
    por_tipo = {}
    for b in qs:
        por_tipo[b.tipo or "(sin tipo)"] = por_tipo.get(b.tipo or "(sin tipo)", 0) + 1
    resumen.append(["Métrica", "Valor"])
    resumen.append(["Total beneficiarios activos", total])
    resumen.append(["Filtro aplicado", tipo_filter or "Todos"])
    resumen.append(["Fecha descarga", datetime.now().strftime("%Y-%m-%d %H:%M")])
    resumen.append([])
    resumen.append(["Tipo", "Cantidad"])
    for t, c in sorted(por_tipo.items()):
        resumen.append([t, c])
    # Estilo resumen
    for row in resumen.iter_rows(min_row=1, max_row=1, max_col=2):
        for cell in row:
            cell.font = header_font; cell.fill = header_fill; cell.alignment = header_align
    resumen.column_dimensions["A"].width = 30
    resumen.column_dimensions["B"].width = 20

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    nombre = (
        f"beneficiarios_{tipo_filter.lower() or 'todos'}_"
        f"{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    )
    response = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{nombre}"'
    return response


@jwt_or_session_required
@modulo_required('org_admin')
def beneficiarios_exportar_csv(request):
    """Descarga CSV de beneficiarios. Respeta el filtro `?tipo=X` actual.

    Sirve para análisis externo (Excel, Power BI) y reportes manuales.
    """
    import csv
    from django.http import HttpResponse
    from datetime import datetime

    qs = (
        Beneficiario.objects
        .select_related("persona", "proveedor", "organizacion", "tipo_documento")
        .filter(activo=True)
    )
    tipo_filter = (request.GET.get("tipo") or "").strip().upper()
    if tipo_filter in {"PERSONA", "ORGANIZACION", "PROVEEDOR"}:
        qs = qs.filter(tipo=tipo_filter)
    qs = qs.order_by("tipo", "nombre_legal")

    nombre_archivo = (
        f"beneficiarios_{tipo_filter.lower() or 'todos'}_"
        f"{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
    )
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{nombre_archivo}"'
    # BOM para que Excel abra UTF-8 correctamente con tildes.
    response.write("﻿")

    writer = csv.writer(response, delimiter=";", quoting=csv.QUOTE_MINIMAL)
    writer.writerow([
        "ID", "Tipo", "Tipo Documento", "Número Documento", "Nombre Legal",
        "Correo", "Teléfono", "Dirección",
        "Persona ID", "Persona Nombre",
        "Organización ID", "Organización Nombre", "Organización NIT",
        "Proveedor ID", "Proveedor Nombre",
        "Activo",
    ])
    for b in qs.iterator(chunk_size=500):
        persona_nombre = ""
        if b.persona:
            partes = [b.persona.nombre1 or "", b.persona.nombre2 or "",
                      b.persona.apellido1 or "", b.persona.apellido2 or ""]
            persona_nombre = " ".join(p for p in partes if p).strip()
        writer.writerow([
            b.id,
            b.tipo or "",
            getattr(b.tipo_documento, "nombre", "") or "",
            b.numero_documento or "",
            b.nombre_legal or "",
            b.correo or "",
            b.telefono or "",
            b.direccion or "",
            b.persona_id or "",
            persona_nombre,
            b.organizacion_id or "",
            getattr(b.organizacion, "nombre", "") or "",
            getattr(b.organizacion, "nit", "") or "",
            b.proveedor_id or "",
            getattr(b.proveedor, "nombre", "") or "",
            "Sí" if b.activo else "No",
        ])
    return response


@login_required
@modulo_required('org_admin')
def beneficiario_nuevo(request):
    """Migrado a Angular: alta de beneficiario (form inline)."""
    return redirect("/app/admin/org")
