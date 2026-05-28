"""Reporte consolidado del curso (PR-D Curso Docente).

Cruza asistencia (PR-B) + notas (PR-C) en una sola estructura por
participante, lista para renderizar HTML, exportar a Excel/PDF
(reportlab/openpyxl) o servir como JSON desde el endpoint DRF.

Sin DDL: consume tablas ya vivas (`participante_evento`,
`asistencia_clase`, `clase`, `evaluacion_participante`).
"""
from dataclasses import dataclass
from decimal import Decimal

from apps.login.models.curso_sesiones import (
    AsistenciaClase, Clase, EvaluacionParticipante,
)
from apps.login.services.curso_sesiones import inscritos_de_curso


@dataclass(frozen=True)
class FilaReporte:
    participante_id: int
    persona_nombre: str
    documento: str
    asistencias: int
    inasistencias: int
    total_marcas: int
    pct_asistencia: float | None  # None si no hay marcas
    notas: list[str]  # ["4.50 (Parcial 1) 2026-04-20", ...]
    promedio: Decimal | None
    aprobado: bool | None  # según UMBRAL_APROBACION


UMBRAL_APROBACION = Decimal('3.0')  # SED Bogotá clásico
UMBRAL_ASISTENCIA_PCT = 80.0  # asistencia mínima sugerida


def _doc_de_persona(persona) -> str:
    """Resuelve el documento sin asumir que la columna existe."""
    pd = getattr(persona, 'persona_documento', None)
    if pd is not None:
        num = getattr(pd, 'numero_documento', None)
        if num:
            return str(num)
    # Fallback: campo `documento` plano si existe
    doc = getattr(persona, 'documento', None) or ''
    return str(doc).strip()


def reporte_consolidado(evento_id: int) -> list[FilaReporte]:
    """Fila por participante con asistencia + notas + promedio.

    Devuelve lista ordenada por apellido. Si un participante no
    tiene marcas ni notas, igual aparece (con vacíos).
    """
    inscritos = list(inscritos_de_curso(evento_id))

    # Pre-agrega marcas por participante (una sola query)
    marcas = list(AsistenciaClase.objects
                  .filter(clase__evento_id=evento_id)
                  .values('participante_id', 'asistencia'))
    asist_por_p: dict[int, dict[str, int]] = {}
    for m in marcas:
        d = asist_por_p.setdefault(m['participante_id'], {'p': 0, 'a': 0})
        if m['asistencia']:
            d['p'] += 1
        else:
            d['a'] += 1

    # Notas por participante
    notas_db = list(EvaluacionParticipante.objects
                    .filter(evento_id=evento_id)
                    .order_by('participante_id', 'fecha_evaluacion', 'id')
                    .values('participante_id', 'resultado',
                            'observaciones', 'fecha_evaluacion'))
    notas_por_p: dict[int, list[dict]] = {}
    for n in notas_db:
        notas_por_p.setdefault(n['participante_id'], []).append(n)

    filas: list[FilaReporte] = []
    for pe in inscritos:
        p = pe.participante
        persona = p.persona
        ap = asist_por_p.get(p.id, {'p': 0, 'a': 0})
        presentes = ap['p']
        ausentes = ap['a']
        total = presentes + ausentes
        pct = (presentes * 100.0 / total) if total else None

        nlist = notas_por_p.get(p.id, [])
        # Formato corto: "4.50 (Parcial 1) 2026-04-20"
        notas_fmt = []
        suma = Decimal('0')
        n_validos = 0
        for n in nlist:
            try:
                v = Decimal(n['resultado'])
            except Exception:
                continue
            suma += v
            n_validos += 1
            etq = f" ({n['observaciones']})" if n['observaciones'] else ''
            fch = f" {n['fecha_evaluacion']:%Y-%m-%d}" if n['fecha_evaluacion'] else ''
            notas_fmt.append(f"{v}{etq}{fch}".strip())

        promedio = (suma / n_validos).quantize(Decimal('0.01')) if n_validos else None
        aprobado = None
        if promedio is not None:
            aprobado = promedio >= UMBRAL_APROBACION

        nombre = f'{persona.nombre1 or ""} {persona.apellido1 or ""}'.strip()

        filas.append(FilaReporte(
            participante_id=p.id,
            persona_nombre=nombre or f'Participante #{p.id}',
            documento=_doc_de_persona(persona),
            asistencias=presentes,
            inasistencias=ausentes,
            total_marcas=total,
            pct_asistencia=pct,
            notas=notas_fmt,
            promedio=promedio,
            aprobado=aprobado,
        ))

    return filas


def reporte_a_dict(evento_id: int) -> list[dict]:
    """Versión JSON-serializable del reporte (para DRF endpoint)."""
    out = []
    for f in reporte_consolidado(evento_id):
        out.append({
            'participante_id': f.participante_id,
            'persona_nombre': f.persona_nombre,
            'documento': f.documento,
            'asistencias': f.asistencias,
            'inasistencias': f.inasistencias,
            'total_marcas': f.total_marcas,
            'pct_asistencia': round(f.pct_asistencia, 1) if f.pct_asistencia is not None else None,
            'notas': f.notas,
            'promedio': str(f.promedio) if f.promedio is not None else None,
            'aprobado': f.aprobado,
        })
    return out


# ─────────────────────────────────────────────────────────────────────────
# Exports — Excel (openpyxl) y PDF (reportlab)
# ─────────────────────────────────────────────────────────────────────────


def exportar_excel(evento) -> bytes:
    """Genera reporte XLSX del curso.

    Devuelve bytes listos para HttpResponse. Una hoja "Reporte"
    con header + filas por participante.
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    filas = reporte_consolidado(evento.id)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Reporte'

    # Título
    ws['A1'] = f'Reporte de curso: {evento.nombre or evento.id}'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:I1')
    if evento.fecha_inicio:
        ws['A2'] = f'Fecha inicio: {evento.fecha_inicio:%Y-%m-%d}'
        ws['A2'].font = Font(italic=True, color='666666')

    # Header
    headers = ['#', 'Documento', 'Participante',
               'Presentes', 'Ausentes', 'Total marcas',
               '% asistencia', 'Notas', 'Promedio', 'Aprobado']
    header_fill = PatternFill('solid', fgColor='305496')
    header_font = Font(bold=True, color='FFFFFF')
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=col, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Filas
    for i, f in enumerate(filas, start=1):
        r = 4 + i
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=f.documento or '—')
        ws.cell(row=r, column=3, value=f.persona_nombre)
        ws.cell(row=r, column=4, value=f.asistencias)
        ws.cell(row=r, column=5, value=f.inasistencias)
        ws.cell(row=r, column=6, value=f.total_marcas)
        if f.pct_asistencia is not None:
            ws.cell(row=r, column=7, value=round(f.pct_asistencia, 1))
        ws.cell(row=r, column=8, value=' | '.join(f.notas) if f.notas else '—')
        if f.promedio is not None:
            ws.cell(row=r, column=9, value=float(f.promedio))
        if f.aprobado is True:
            ws.cell(row=r, column=10, value='SÍ')
        elif f.aprobado is False:
            ws.cell(row=r, column=10, value='NO')
        else:
            ws.cell(row=r, column=10, value='—')

    # Anchos
    widths = [4, 14, 32, 10, 10, 12, 12, 38, 10, 10]
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + col_idx)].width = w

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def exportar_pdf(evento) -> bytes:
    """Genera reporte PDF del curso usando reportlab Platypus."""
    import io
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
    )

    filas = reporte_consolidado(evento.id)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(letter),
        leftMargin=1.5 * cm, rightMargin=1.5 * cm,
        topMargin=1.5 * cm, bottomMargin=1.5 * cm,
    )

    styles = getSampleStyleSheet()
    elems = []
    titulo = f"Reporte del curso: {evento.nombre or evento.id}"
    elems.append(Paragraph(titulo, styles['Title']))
    if evento.fecha_inicio:
        elems.append(Paragraph(
            f"<i>Fecha inicio: {evento.fecha_inicio:%Y-%m-%d}</i>",
            styles['Normal'],
        ))
    elems.append(Spacer(1, 12))

    # Datos de la tabla
    data = [['#', 'Doc.', 'Participante', 'Pres.', 'Aus.', '% Asist.',
             'Notas', 'Prom.', 'Apr.']]
    for i, f in enumerate(filas, start=1):
        data.append([
            i,
            f.documento or '—',
            f.persona_nombre[:32],
            f.asistencias,
            f.inasistencias,
            f'{f.pct_asistencia:.1f}%' if f.pct_asistencia is not None else '—',
            (' | '.join(f.notas))[:60] if f.notas else '—',
            str(f.promedio) if f.promedio is not None else '—',
            ('SÍ' if f.aprobado else 'NO') if f.aprobado is not None else '—',
        ])

    table = Table(data, repeatRows=1,
                  colWidths=[1*cm, 2.5*cm, 5.5*cm, 1.3*cm, 1.3*cm,
                             1.8*cm, 7*cm, 1.5*cm, 1.2*cm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#305496')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (3, 1), (5, -1), 'CENTER'),
        ('ALIGN', (7, 1), (8, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1),
            [colors.white, colors.HexColor('#F0F0F0')]),
    ]))
    elems.append(table)
    doc.build(elems)
    return buf.getvalue()
