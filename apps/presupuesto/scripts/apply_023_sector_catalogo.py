"""Aplica el 023: el catálogo de SECTOR, y lo siembra desde la Matriz PDL.

    docker exec innova_k python apps/presupuesto/scripts/apply_023_sector_catalogo.py --seco
    docker exec innova_k python apps/presupuesto/scripts/apply_023_sector_catalogo.py
    docker exec innova_k python apps/presupuesto/scripts/apply_023_sector_catalogo.py --rollback

⚠️ REQUIERE APROBACIÓN EXPLÍCITA DE ALEX (Constitución VII) y backup < 24 h.

**SECO POR DEFECTO NO**: acá el default es aplicar, como los apply_0NN
hermanos. Para ensayar sin escribir, `--seco` corre todo dentro de una
transacción y hace ROLLBACK al final, imprimiendo lo mismo que imprimiría.

Es ADITIVO (dos tablas nuevas + una columna nullable en `metas`) e IDEMPOTENTE:
correrlo dos veces no duplica sectores ni alias, y el backfill vuelve a
calcular lo mismo.

QUÉ SIEMBRA, Y DE DÓNDE
-----------------------
El catálogo NO está escrito a mano en este archivo: se lee de la hoja
«Programacion PDL 2025 - 2028» de la matriz. Si la ALK agrega un sector en el
próximo corte, entra solo. Los 13 valores medidos el 2026-09-03 son:

    GOBIERNO · SEGURIDAD, CONVIVENCIA Y JUSTICIA · CULTURA, RECREACIÓN Y
    DEPORTE · INTEGRACIÓN SOCIAL · AMBIENTE · AMBIENTE/HÁBITAT · GESTIÓN
    PÚBLICA · SALUD · EDUCACIÓN · DESARROLLO ECONÓMICO, INDUSTRIA Y TURISMO ·
    MUJERES/INTEGRACIÓN SOCIAL · MOVILIDAD · MUJERES

El backfill de `metas.sector_id` va por la llave ESTABLE
`(proyecto_codigo, codind)` y jamás por el texto de `metas.sector`. Esa es la
decisión que hace que el gráfico deje de estar partido: el texto miente en 23
de 78 filas; la llave no.

Los ALIAS se siembran solo donde la matriz demuestra que son la misma cosa, y
el propio script RECHAZA cualquier alias ambiguo (ver `--seco`): así
'Infraestructura' —que mapea a MOVILIDAD y a CULTURA a la vez— no entra ni por
descuido.
"""
import argparse
import os
import sys
import unicodedata
from pathlib import Path

import django

BASE = Path(__file__).resolve().parents[3]
sys.path.insert(0, str(BASE))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "core.settings")
django.setup()

from django.db import connection, transaction  # noqa: E402

AQUI = Path(__file__).resolve().parent
TABLAS = ("presu_sector", "presu_sector_alias")

# El archivo de la matriz. No se versiona (está en .gitignore): pertenece a su
# carga, no al repo, que es público.
MATRIZ = BASE / "Matriz de seguimiento PDL 2025-2028.xlsx"
HOJA_PROG = "Programacion PDL 2025 - 2028"

# Columnas de la hoja «Programacion», por POSICIÓN 0-based. Se validan contra
# el encabezado antes de leer: si la ALK mueve una columna, el script aborta en
# vez de cargar la columna de al lado.
COL_SECTOR = 2      # C · Sector
COL_INDICADOR = 3   # D · No. Indicador
COL_PROYECTO = 8    # I · Cód. Proyecto de Inversión SEGPLAN
ENCABEZADOS = {
    COL_SECTOR: "Sector",
    COL_INDICADOR: "No. Indicador",
    COL_PROYECTO: "Cód. Proyecto de Inversión SEGPLAN",
}


def norm_texto(v):
    """Mayúsculas, sin tildes, sin espacios dobles. UNA sola implementación.

    La usa el sembrado y la tiene que usar la ingesta: si se escriben dos, se
    separan y el mismo texto empieza a resolver distinto según por dónde entró.
    """
    if v is None:
        return ""
    s = unicodedata.normalize("NFKD", str(v))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return " ".join(s.upper().split())


def norm_codigo(v):
    """Quita ceros a la izquierda. La matriz guarda 2377; innovaK, 0002377."""
    s = "" if v is None else str(v).strip()
    if s.endswith(".0"):          # openpyxl devuelve 2377.0 en celdas numéricas
        s = s[:-2]
    return s.lstrip("0") or "0"


def leer_matriz():
    """Devuelve (sectores_ordenados, {(proyecto, indicador): sector})."""
    try:
        import openpyxl
    except ImportError:
        sys.exit("error: falta openpyxl en el contenedor.")
    if not MATRIZ.exists():
        sys.exit(f"error: no encuentro la matriz en {MATRIZ}")

    wb = openpyxl.load_workbook(MATRIZ, data_only=True, read_only=True)
    if HOJA_PROG not in wb.sheetnames:
        sys.exit(f"error: la hoja «{HOJA_PROG}» no está. Hay: {wb.sheetnames}")
    ws = wb[HOJA_PROG]

    filas = ws.iter_rows(values_only=True)
    cabecera = next(filas)
    for idx, esperado in ENCABEZADOS.items():
        visto = norm_texto(cabecera[idx] if idx < len(cabecera) else None)
        if visto != norm_texto(esperado):
            sys.exit(
                f"error: la columna {idx} debería ser «{esperado}» y dice "
                f"«{cabecera[idx]}». La matriz cambió de forma: revisar "
                f"docs/operacion/matriz_pdl_mapeo.md antes de seguir."
            )

    por_llave, orden = {}, []
    for fila in filas:
        if not fila or all(c is None for c in fila):
            continue
        sector = (fila[COL_SECTOR] or "").strip() if fila[COL_SECTOR] else ""
        if not sector:
            continue
        if sector not in orden:
            orden.append(sector)
        proy = norm_codigo(fila[COL_PROYECTO])
        ind = norm_codigo(fila[COL_INDICADOR])
        if proy and ind:
            por_llave[(proy, ind)] = sector
    return orden, por_llave


def alias_desde_datos(cur, por_llave):
    """Deduce los alias mirando qué dice la matriz de cada meta.

    Devuelve (aceptados, rechazados). Un valor de `metas.sector` que la matriz
    resuelve a DOS sectores distintos se rechaza: no es un alias, y meterlo
    mentiría de uno de los dos lados.
    """
    cur.execute("SELECT proyecto_codigo, codind, sector FROM metas "
                "WHERE sector IS NOT NULL AND sector <> ''")
    candidatos = {}
    for proy, ind, texto in cur.fetchall():
        oficial = por_llave.get((norm_codigo(proy), norm_codigo(ind)))
        if not oficial:
            continue
        if norm_texto(texto) == norm_texto(oficial):
            continue                      # ya está en el vocabulario oficial
        candidatos.setdefault(texto, set()).add(oficial)

    aceptados, ambiguos, dudosos = {}, {}, {}
    for texto, cuales in candidatos.items():
        if len(cuales) > 1:
            ambiguos[texto] = sorted(cuales)
            continue
        oficial = next(iter(cuales))
        # No basta con que sea inequívoco: tiene que ser PLAUSIBLE.
        #
        # «Sin ambigüedad» solo dice que las metas con ese texto cayeron todas
        # en el mismo sector. Con UNA sola meta eso no prueba nada: 'CPS y
        # Planta' resolvía limpio a GOBIERNO y no es un sector, es un tipo de
        # contratación que alguien escribió en la columna equivocada.
        #
        # La prueba que sí separa los dos casos: un alias de verdad aparece
        # DENTRO del nombre oficial. 'Deporte' está en 'CULTURA, RECREACIÓN Y
        # DEPORTE' y 'Seguridad' en 'SEGURIDAD, CONVIVENCIA Y JUSTICIA'; 'CPS
        # y Planta' no está en 'GOBIERNO'. Lo que no pasa la prueba no se
        # descarta: se lista para que lo confirme una persona.
        if norm_texto(texto) in norm_texto(oficial):
            aceptados[texto] = oficial
        else:
            dudosos[texto] = oficial
    return aceptados, ambiguos, dudosos


def sembrar(cur, seco):
    orden, por_llave = leer_matriz()
    print(f"  matriz: {len(orden)} sectores, {len(por_llave)} pares (proyecto, indicador)")

    # ── catálogo ──
    nuevos = 0
    for nombre in orden:
        cur.execute(
            "INSERT INTO presu_sector (nombre_oficial) VALUES (%s) "
            "ON CONFLICT (nombre_oficial) DO NOTHING", [nombre])
        nuevos += cur.rowcount
    print(f"  presu_sector: +{nuevos} nuevos de {len(orden)}")

    cur.execute("SELECT nombre_oficial, id FROM presu_sector")
    id_por_nombre = {n: i for n, i in cur.fetchall()}

    # ── alias, deducidos y filtrados ──
    aceptados, ambiguos, dudosos = alias_desde_datos(cur, por_llave)
    n_alias = 0
    for texto, oficial in sorted(aceptados.items()):
        sid = id_por_nombre.get(oficial)
        if sid is None:
            continue
        cur.execute(
            "INSERT INTO presu_sector_alias (sector_id, alias, alias_norm, origen) "
            "VALUES (%s, %s, %s, 'innovak_interno') "
            "ON CONFLICT (alias_norm) DO NOTHING",
            [sid, texto, norm_texto(texto)])
        n_alias += cur.rowcount
        print(f"    alias  {texto!r} → {oficial!r}")
    print(f"  presu_sector_alias: +{n_alias}")

    if ambiguos:
        print("  FUERA por ambiguos — mapean a dos sectores, no son alias:")
        for texto, cuales in sorted(ambiguos.items()):
            print(f"    {texto!r} → {cuales}")
    if dudosos:
        print("  FUERA por no plausibles — inequívocos pero el texto no está")
        print("  en el nombre oficial. Confirmar con la ALK antes de sembrarlos:")
        for texto, oficial in sorted(dudosos.items()):
            print(f"    {texto!r} → {oficial!r} ?")

    # ── backfill de metas.sector_id, en DOS pasos y en este orden ──
    #
    # 1) Por la llave estable (proyecto, indicador) contra la matriz. Es la
    #    autoridad: el texto de `metas.sector` miente en 23 de 78 filas.
    # 2) Solo para lo que la matriz NO cubre, por el TEXTO normalizado, contra
    #    el nombre oficial primero y contra los alias después.
    #
    # El paso 2 no es un atajo del 1: es para las filas que la matriz no puede
    # resolver porque no tienen con qué cruzarse. Sin él, la meta 8
    # ('Educación', «becas de educación», sin proyecto_codigo ni codind) caía
    # en «Sin sector» y se llevaba TODO el avance de Educación al primer lugar
    # del gráfico como una barra anónima — cambiando un defecto por otro.
    por_norma = {norm_texto(n): i for n, i in id_por_nombre.items()}
    cur.execute("SELECT alias_norm, sector_id FROM presu_sector_alias")
    por_norma.update({a: s for a, s in cur.fetchall()})

    cur.execute("SELECT codigo, proyecto_codigo, codind, sector FROM metas")
    filas = cur.fetchall()
    por_matriz, por_texto, sin_resolver = 0, [], []
    for cod, proy, ind, texto in filas:
        oficial = por_llave.get((norm_codigo(proy), norm_codigo(ind)))
        sid = id_por_nombre.get(oficial) if oficial else por_norma.get(norm_texto(texto))
        if sid is None:
            sin_resolver.append((cod, texto))
            continue
        cur.execute("UPDATE metas SET sector_id = %s WHERE codigo = %s "
                    "AND (sector_id IS DISTINCT FROM %s)", [sid, cod, sid])
        if oficial:
            por_matriz += cur.rowcount
        else:
            por_texto.append((cod, texto))

    print(f"  metas.sector_id: {por_matriz} por la matriz · "
          f"{len(por_texto)} por texto {[c for c, _ in por_texto]}")
    if sin_resolver:
        print("  sin sector (no cruzan la matriz y el texto no es un sector):")
        for cod, texto in sin_resolver:
            print(f"    meta {cod}: {texto!r}")

    if seco:
        print("\n  --seco: se hace ROLLBACK, nada quedó escrito.")


def estado(cur):
    for tabla in TABLAS:
        cur.execute("""SELECT EXISTS (SELECT 1 FROM information_schema.tables
                       WHERE table_schema='public' AND table_name=%s)""", [tabla])
        if not cur.fetchone()[0]:
            print(f"  {tabla:24s} —")
            continue
        cur.execute(f'SELECT count(*) FROM "{tabla}"')
        print(f"  {tabla:24s} {cur.fetchone()[0]} fila(s)")
    cur.execute("""SELECT EXISTS (SELECT 1 FROM information_schema.columns
                   WHERE table_name='metas' AND column_name='sector_id')""")
    if cur.fetchone()[0]:
        cur.execute("SELECT count(sector_id), count(*) FROM metas")
        con, total = cur.fetchone()
        print(f"  metas.sector_id          {con} de {total} metas")
    else:
        print("  metas.sector_id          —")


def correr_sql(cur, nombre):
    """Ejecuta un .sql de este directorio SIN su control de transacción.

    Los `.sql` traen su propio `BEGIN;`/`COMMIT;` para poder correrlos a mano
    con psql, que es como se usan los hermanos. Pero acá el guion necesita
    mandar él en la transacción: si el `COMMIT;` del archivo se ejecutara
    dentro del `atomic()`, cerraría la transacción a mitad de camino y
    `--seco` escribiría de verdad — justo lo que promete no hacer.

    Solo se quitan las líneas que son EXACTAMENTE `BEGIN;` o `COMMIT;`. El
    `BEGIN` del bloque `DO $$ … END $$;` no lleva punto y coma y no se toca.
    """
    sql = (AQUI / nombre).read_text(encoding="utf-8")
    limpio = "\n".join(
        linea for linea in sql.splitlines()
        if linea.strip().upper() not in ("BEGIN;", "COMMIT;")
    )
    cur.execute(limpio)


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--rollback", action="store_true", help="deshace el 023")
    ap.add_argument("--seco", action="store_true",
                    help="ensaya todo y hace ROLLBACK: no escribe nada")
    args = ap.parse_args()

    with connection.cursor() as cur:
        print("ANTES:")
        estado(cur)

    if args.rollback:
        with transaction.atomic(), connection.cursor() as cur:
            correr_sql(cur, "rollback_023_sector_catalogo.sql")
        with connection.cursor() as cur:
            print("\nDESPUÉS del rollback:")
            estado(cur)
        return

    try:
        with transaction.atomic():
            with connection.cursor() as cur:
                print("\nAPLICANDO 023…")
                correr_sql(cur, "023_sector_catalogo.sql")
                sembrar(cur, args.seco)
            if args.seco:
                raise _Seco()
    except _Seco:
        pass

    with connection.cursor() as cur:
        print("\nDESPUÉS:")
        estado(cur)


class _Seco(Exception):
    """Señal interna para que `--seco` revierta la transacción."""


if __name__ == "__main__":
    main()
