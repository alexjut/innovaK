"""Aplica el 024: Objetivo Estratégico y Programa como tablas, sembrados
desde la Matriz PDL.

    docker exec innova_k python apps/presupuesto/scripts/apply_024_objetivo_programa.py --seco
    docker exec innova_k python apps/presupuesto/scripts/apply_024_objetivo_programa.py
    docker exec innova_k python apps/presupuesto/scripts/apply_024_objetivo_programa.py --rollback

⚠️ REQUIERE APROBACIÓN EXPLÍCITA DE ALEX (Constitución VII) y backup < 24 h.

Mismo patrón que `apply_023_sector_catalogo.py`: `--seco` corre todo dentro de
una transacción y hace ROLLBACK, aditivo e idempotente.

QUÉ SIEMBRA
-----------
Nada está escrito a mano: los 5 objetivos y los 22 programas se leen de la hoja
«Seguimiento», y el código sale del propio nombre («3 - Bogotá confía en su
potencial» → código 3). Si la ALK agrega un programa, entra solo.

El backfill de `metas.programa_id` va por la llave ESTABLE
`(proyecto_codigo, codind)`, igual que el sector — nunca por `metas.codprog`,
que es texto backfilleado y puede haber quedado viejo por la misma razón que
el sector: el importador no pisa lo que ya está escrito.

Y no se limita a cargar: COMPARA lo cargado contra `metas.codprog` y reporta
las divergencias, que es la única forma de saber si el texto que hoy leen las
pantallas dice lo mismo que la matriz.
"""
import argparse
import os
import re
import sys
import unicodedata
from pathlib import Path

import django

BASE = Path(__file__).resolve().parents[3]
sys.path.insert(0, str(BASE))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "core.settings")
django.setup()

from django.db import connection, transaction  # noqa: E402

AQUI = Path(__file__).resolve().parent
TABLAS = ("presu_objetivo_estrategico", "presu_programa")

MATRIZ = BASE / "Matriz de seguimiento PDL 2025-2028.xlsx"
HOJA_SEG = "Seguimiento"

# Columnas de la hoja «Seguimiento», 0-based, validadas contra el encabezado.
COL_OBJETIVO = 0    # A · Objetivo  Estrategico
COL_PROGRAMA = 1    # B · Programa
COL_PROYECTO = 6    # G · N° Proyecto de inversión
COL_INDICADOR = 9   # J · Codigo indicador
ENCABEZADOS = {
    COL_OBJETIVO: "Objetivo Estrategico",
    COL_PROGRAMA: "Programa",
    COL_PROYECTO: "N° Proyecto de inversión",
    COL_INDICADOR: "Codigo indicador",
}

# «3 - Bogotá confía en su potencial» → (3, 'Bogotá confía en su potencial')
PATRON = re.compile(r"^\s*(\d+)\s*-\s*(.+?)\s*$")


def norm_texto(v):
    if v is None:
        return ""
    s = unicodedata.normalize("NFKD", str(v))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return " ".join(s.upper().split())


def norm_codigo(v):
    s = "" if v is None else str(v).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s.lstrip("0") or "0"


def partir(texto):
    """«N - nombre» → (N, nombre). Aborta si no calza: inventar un código
    sería peor que parar."""
    m = PATRON.match(" ".join(str(texto).split()))
    if not m:
        sys.exit(f"error: «{texto}» no tiene la forma «N - nombre». "
                 f"La matriz cambió de convención: revisar antes de seguir.")
    return int(m.group(1)), m.group(2)


def leer_matriz():
    try:
        import openpyxl
    except ImportError:
        sys.exit("error: falta openpyxl en el contenedor.")
    if not MATRIZ.exists():
        sys.exit(f"error: no encuentro la matriz en {MATRIZ}")

    wb = openpyxl.load_workbook(MATRIZ, data_only=True, read_only=True)
    if HOJA_SEG not in wb.sheetnames:
        sys.exit(f"error: falta la hoja «{HOJA_SEG}». Hay: {wb.sheetnames}")
    filas = wb[HOJA_SEG].iter_rows(values_only=True)

    cabecera = next(filas)
    for idx, esperado in ENCABEZADOS.items():
        visto = norm_texto(cabecera[idx] if idx < len(cabecera) else None)
        if visto != norm_texto(esperado):
            sys.exit(f"error: la columna {idx} debería ser «{esperado}» y dice "
                     f"«{cabecera[idx]}». Revisar matriz_pdl_mapeo.md.")

    objetivos, programas, por_llave, por_proyecto = {}, {}, {}, {}
    for fila in filas:
        if not fila or all(c is None for c in fila):
            continue
        if not fila[COL_OBJETIVO] or not fila[COL_PROGRAMA]:
            continue
        cod_o, nom_o = partir(fila[COL_OBJETIVO])
        cod_p, nom_p = partir(fila[COL_PROGRAMA])
        objetivos[cod_o] = nom_o

        previo = programas.get(cod_p)
        if previo and previo[1] != cod_o:
            sys.exit(f"error: el programa {cod_p} cuelga de los objetivos "
                     f"{previo[1]} y {cod_o}. El modelo asume UN padre: "
                     f"si esto pasa, hace falta tabla puente, no un parche.")
        programas[cod_p] = (nom_p, cod_o)

        proy, ind = norm_codigo(fila[COL_PROYECTO]), norm_codigo(fila[COL_INDICADOR])
        if proy and ind:
            por_llave[(proy, ind)] = cod_p
        if proy:
            por_proyecto.setdefault(proy, set()).add(cod_p)

    # `proyecto → programa` es una FUNCIÓN en la matriz: verificado el
    # 2026-09-03, 0 de 30 proyectos tienen más de un programa. Eso habilita el
    # respaldo de abajo. Igual se comprueba acá y no se da por sentado: el día
    # que un proyecto tenga dos, ese proyecto se cae del respaldo en vez de
    # elegir uno al azar.
    proy_a_programa = {p: next(iter(s)) for p, s in por_proyecto.items() if len(s) == 1}
    return objetivos, programas, por_llave, proy_a_programa


def sembrar(cur, seco):
    objetivos, programas, por_llave, proy_a_programa = leer_matriz()
    print(f"  matriz: {len(objetivos)} objetivos · {len(programas)} programas "
          f"· {len(por_llave)} pares (proyecto, indicador)")

    nuevos_o = 0
    for cod, nombre in sorted(objetivos.items()):
        cur.execute("INSERT INTO presu_objetivo_estrategico (codigo, nombre) "
                    "VALUES (%s, %s) ON CONFLICT (codigo) DO NOTHING", [cod, nombre])
        nuevos_o += cur.rowcount
    cur.execute("SELECT codigo, id FROM presu_objetivo_estrategico")
    id_objetivo = dict(cur.fetchall())
    print(f"  presu_objetivo_estrategico: +{nuevos_o} de {len(objetivos)}")

    nuevos_p = 0
    for cod, (nombre, cod_o) in sorted(programas.items()):
        cur.execute("INSERT INTO presu_programa (codigo, nombre, objetivo_id) "
                    "VALUES (%s, %s, %s) ON CONFLICT (codigo) DO NOTHING",
                    [cod, nombre, id_objetivo[cod_o]])
        nuevos_p += cur.rowcount
    cur.execute("SELECT codigo, id FROM presu_programa")
    id_programa = dict(cur.fetchall())
    print(f"  presu_programa: +{nuevos_p} de {len(programas)}")

    # ── backfill en DOS pasos, y comparación con el texto viejo ──
    #
    # 1) Por la llave estable (proyecto_codigo, codind) contra la matriz.
    # 2) Para lo que el paso 1 no cubre, por `meta_proyecto` → proyecto.
    #
    # El paso 2 existe por las metas AGRUPADAS. La meta 8 dice «Impactar 1400
    # jóvenes con becas de educación» y en la matriz son DOS filas de 700 del
    # proyecto 2377 (indicadores 51 y 52): innovaK las tiene colapsadas en una,
    # así que no tiene `codind` y por la llave estable no cruza nunca.
    #
    # No es adivinar: `proyecto → programa` es una función en la matriz, y
    # `leer_matriz` solo deja en `proy_a_programa` los proyectos donde de veras
    # lo es. Un proyecto con dos programas queda fuera y su meta se va en NULL.
    cur.execute("""SELECT m.codigo, m.proyecto_codigo, m.codind, m.codprog, p.codigo
                   FROM metas m
                   LEFT JOIN meta_proyecto mp ON mp.meta_id = m.codigo
                   LEFT JOIN proyecto p ON p.id = mp.proyecto_id""")
    tocadas, por_respaldo, sin_cruce, divergen = 0, [], [], []
    for cod_meta, proy, ind, codprog, proy_vinculado in cur.fetchall():
        cod_p = por_llave.get((norm_codigo(proy), norm_codigo(ind)))
        if cod_p is None and proy_vinculado:
            cod_p = proy_a_programa.get(norm_codigo(proy_vinculado))
            if cod_p is not None:
                por_respaldo.append((cod_meta, norm_codigo(proy_vinculado)))
        if cod_p is None:
            sin_cruce.append(cod_meta)
            continue
        pid = id_programa[cod_p]
        cur.execute("UPDATE metas SET programa_id = %s WHERE codigo = %s "
                    "AND (programa_id IS DISTINCT FROM %s)", [pid, cod_meta, pid])
        tocadas += cur.rowcount
        if codprog is not None and int(codprog) != cod_p:
            divergen.append((cod_meta, int(codprog), cod_p))

    print(f"  metas.programa_id: {tocadas} actualizadas · "
          f"{len(sin_cruce)} sin cruce {sin_cruce}")
    if por_respaldo:
        print(f"  por respaldo meta_proyecto ({len(por_respaldo)}) "
              f"— metas agrupadas, sin codind propio:")
        for cod_meta, proy in por_respaldo:
            print(f"    meta {cod_meta} → proyecto {proy}")
    if divergen:
        print(f"  DIVERGEN de metas.codprog ({len(divergen)}) "
              f"— el texto viejo decía otra cosa que la matriz:")
        for cod_meta, viejo, nuevo in divergen[:15]:
            print(f"    meta {cod_meta}: codprog={viejo} → programa {nuevo}")
    else:
        print("  metas.codprog coincide con la matriz en todas las que cruzan.")

    if seco:
        print("\n  --seco: se hace ROLLBACK, nada quedó escrito.")


def estado(cur):
    for tabla in TABLAS:
        cur.execute("""SELECT EXISTS (SELECT 1 FROM information_schema.tables
                       WHERE table_schema='public' AND table_name=%s)""", [tabla])
        if not cur.fetchone()[0]:
            print(f"  {tabla:30s} —")
            continue
        cur.execute(f'SELECT count(*) FROM "{tabla}"')
        print(f"  {tabla:30s} {cur.fetchone()[0]} fila(s)")
    cur.execute("""SELECT EXISTS (SELECT 1 FROM information_schema.columns
                   WHERE table_name='metas' AND column_name='programa_id')""")
    if cur.fetchone()[0]:
        cur.execute("SELECT count(programa_id), count(*) FROM metas")
        con, total = cur.fetchone()
        print(f"  metas.programa_id              {con} de {total} metas")
    else:
        print("  metas.programa_id              —")


def correr_sql(cur, nombre):
    """Ejecuta el .sql sin su control de transacción: acá manda el guion.

    Los `.sql` traen `BEGIN;`/`COMMIT;` para poder correrlos a mano con psql.
    Si ese `COMMIT;` se ejecutara dentro del `atomic()`, cerraría la
    transacción a mitad y `--seco` escribiría de verdad. El `BEGIN` del bloque
    `DO $$ … END $$;` no lleva punto y coma y no se toca.
    """
    sql = (AQUI / nombre).read_text(encoding="utf-8")
    cur.execute("\n".join(
        l for l in sql.splitlines() if l.strip().upper() not in ("BEGIN;", "COMMIT;")))


class _Seco(Exception):
    """Señal interna para que `--seco` revierta la transacción."""


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--rollback", action="store_true")
    ap.add_argument("--seco", action="store_true",
                    help="ensaya todo y hace ROLLBACK: no escribe nada")
    args = ap.parse_args()

    with connection.cursor() as cur:
        print("ANTES:")
        estado(cur)

    if args.rollback:
        with transaction.atomic(), connection.cursor() as cur:
            correr_sql(cur, "rollback_024_objetivo_programa.sql")
        with connection.cursor() as cur:
            print("\nDESPUÉS del rollback:")
            estado(cur)
        return

    try:
        with transaction.atomic():
            with connection.cursor() as cur:
                print("\nAPLICANDO 024…")
                correr_sql(cur, "024_objetivo_programa.sql")
                sembrar(cur, args.seco)
            if args.seco:
                raise _Seco()
    except _Seco:
        pass

    with connection.cursor() as cur:
        print("\nDESPUÉS:")
        estado(cur)


if __name__ == "__main__":
    main()
