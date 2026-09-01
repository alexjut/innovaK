"""Importa la Matriz de seguimiento PDL que la ALK manda a mano (xlsx).

    docker exec innova_k python manage.py importar_matriz_pdl_alk archivo.xlsx
    docker exec innova_k python manage.py importar_matriz_pdl_alk archivo.xlsx --write --usuario <username>

SECO POR DEFECTO, firmado e IDEMPOTENTE — misma convención que
`sdp_mapear_codigo_meta` y `crear_lineas_plan_faltantes`.

POR QUÉ EXISTE. `sdp_meta_oficial` (RUMBO.md §1.1) es el espejo de la fuente
oficial abierta de Planeación, y esa fuente **está parada desde 2026-02-18**
(escalado a Planeación, sin resolver). La ALK sigue reprogramando metas 2026
igual, y ahora las manda a mano en este Excel porque el canal automático no
las trae.

QUÉ HACE. Arma o completa la cadena viva (`proyecto` → `metas` →
`meta_proyecto` → `presu_indicador_meta_proyecto`) para lo que falte:

- Metas existentes con `codigo_meta` ya enganchado (22 de 24 — ver
  `sdp_mapear_codigo_meta`): se les completan SOLO las columnas SEGPLAN que
  hoy están en NULL (sector, línea, concepto, componente, programa, objetivo,
  codind, nomind). Nunca se pisa un valor ya escrito.
- Indicadores del Excel sin ninguna meta interna enganchada (proyecto nuevo,
  o proyecto existente con un indicador que nunca se cargó): se crea `metas`
  + `meta_proyecto` + el KPI. `meta_magnitud` del KPI toma la magnitud de la
  VIGENCIA ACTUAL (columna "Magnitud Meta Reprogramacion <año>"), siguiendo
  la convención ya viva en los KPI existentes (ver 2377/2780: su magnitud es
  la del año en curso, no el total 2025-2028 — ese total queda en
  `metas.nombre`).
- Proyecto que no existe en absoluto: se crea, con el subgrupo que mejor
  matchea (mapa declarado abajo). Los que no tienen un subgrupo obvio quedan
  marcados "⚠ revisar" en el reporte — se crean igual (para no dejar el
  indicador huérfano) pero con el subgrupo más cercano, para que Alex lo
  corrija con un clic en vez de tener que crear el proyecto él mismo.
- Antes de crear, chequea que el indicador no sea ya un KPI vivo de una META
  AGRUPADA (una sola fila de `metas` que cubre dos indicadores SEGPLAN a la
  vez — pasa en la meta 8 de 2377 — y por eso nunca tiene `codigo_meta`). Si
  hay un KPI activo con la misma magnitud y sin `codigo_meta`, no crea nada:
  reporta y pide revisión. Sin este chequeo la primera corrida de este
  comando duplicó los indicadores 51/52 de 2377 — se detectó con los smoke
  tests y se revirtió a mano.

NO TOCA `meta_magnitud` de un KPI que YA EXISTE, aunque el Excel traiga un
número distinto: la divergencia se REPORTA, no se sobreescribe sola. Cambiar
la meta de un área en curso sin que nadie lo vea es el mismo error que ya le
costó una fila mal enganchada a `sdp_mapear_codigo_meta`.

NO TOCA `sdp_meta_oficial`, aunque esa tabla tenga casi el mismo shape que
este Excel. Se intentó (subir la matriz ahí como una fuente más) y se revirtió
el mismo día: su UNIQUE real es (vigencia, proyecto, indicador) SIN `fuente`,
así que insertar ahí no agrega una fuente en paralelo — PISA la fila oficial
existente. Rompió 10 tests de `apps.dashboard` que suman `total_programado`/
`valor_programado` esperando una sola fuente. Se restauró desde el backup del
día y se sacó el paso. Guardar el presupuesto del Excel en algo persistente
—si hace falta— exige una tabla o columna nueva (DDL), no reusar ésta.

OJO CON LA COLUMNA DE PLATA SI ALGUIEN RETOMA ESO. El primer paso real de
ejecución presupuestal NO es "Presupuesto proyectado PDL" (esa es la meta
aspiracional del cuatrienio) sino **"Apropiación POAI inicial"** — es la que
de verdad se asigna para ejecutar en la vigencia, y puede ser mayor o menor
que la proyectada (medido: 2377/ind.51/2025 proyecta 3.261.800.000 pero apropia
3.751.341.000). La cadena correcta para armar "% de ejecución" es Apropiación
→ Comprometido → Girado, no Proyectado → Comprometido → Girado.
"""
import re

from django.core.management.base import BaseCommand, CommandError

# Sector (Programación PDL) → subgrupo_id, usado SOLO al crear un proyecto que
# no existe todavía. Los que ya existen conservan su subgrupo actual: este
# mapa nunca los toca.
SECTOR_SUBGRUPO_FALLBACK = {
    "AMBIENTE": 10,
    "AMBIENTE/HÁBITAT": 10,
    "CULTURA, RECREACIÓN Y DEPORTE": 1,
    "DESARROLLO ECONÓMICO, INDUSTRIA Y TURISMO": 35,
    "EDUCACIÓN": 8,
    "GESTIÓN PÚBLICA": 9,
    "GOBIERNO": 9,
    "INTEGRACIÓN SOCIAL": 4,
    "MOVILIDAD": 37,
    "MUJERES": 40,
    "MUJERES/INTEGRACIÓN SOCIAL": 40,
    "SALUD": 45,
    "SEGURIDAD, CONVIVENCIA Y JUSTICIA": 38,
}

# Overrides por código de proyecto: el sector solo no alcanza (p.ej. "GOBIERNO"
# cubre desde espacio público hasta comunidades étnicas). (subgrupo_id, dudoso)
PROYECTO_SUBGRUPO_OVERRIDE = {
    2610: (6, False),   # Subsidio tipo C — ya existe, no se usa al crear
    2646: (4, False),   # Kennedy Espacios de Buen Trato → nombre calca el subgrupo
    2684: (43, False),  # Espacios Públicos Seguros → Espacio Público
    2705: (47, False),  # "innovación pública y social fortalecidas" → Innovación
    2733: (3, False),   # organizaciones comunales/sociales → Participación
    2740: (3, True),    # comunidades étnicas (rom, negras, raizales, indígenas):
                          # ningún subgrupo existente nombra "asuntos étnicos".
                          # Participación es el más cercano; revisar con Alex.
    2767: (17, False),  # competencias digitales / servicios TIC → TIC
    2784: (2, False),   # ya existe (Deporte)
    2788: (1, False),   # ya existe (Cultura)
    2790: (37, False),  # ya existe (Infraestructura)
    2793: (37, False),  # "equipamientos culturales" — obra, hermano de 2790
    2551: (37, False),  # espacio público construido — hermano de 2574
    2556: (40, False),  # Mujeres sin Barreras → Mujer
    2729: (38, False),  # sector SEGURIDAD... pese al nombre "…y en Paz"
}


def _sector_a_subgrupo(sector, codigo_proyecto):
    if codigo_proyecto in PROYECTO_SUBGRUPO_OVERRIDE:
        return PROYECTO_SUBGRUPO_OVERRIDE[codigo_proyecto]
    sid = SECTOR_SUBGRUPO_FALLBACK.get((sector or "").strip())
    return (sid, sid is None)


def _num(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _leading_int(s):
    m = re.match(r"\s*(\d+)", str(s or ""))
    return int(m.group(1)) if m else None


# Las cuatro columnas de plata que la hoja «Seguimiento» trae POR VIGENCIA. Se
# emparejan por nombre normalizado y no por posición: los encabezados del Excel
# vienen con espacios de más, saltos de línea en medio y el año pegado al final
# de formas distintas según la columna ("...  2025", "... 2026 ").
CAMPOS_PLATA = {
    "proyectado_pdl":   ("proyectado",),
    "apropiacion_poai": ("apropiaci",),
    "comprometido":     ("comprometido",),
    "girado":           ("girado",),
}
VIGENCIAS = (2025, 2026, 2027, 2028)


def _norm(h):
    return " ".join(str(h or "").split()).lower()


def _columnas_plata(headers):
    """{vigencia: {campo: encabezado_exacto}} a partir de los encabezados.

    Solo mira los que llevan el año: las columnas «Total (2025-2028)» se
    descartan a propósito —acá se guarda el detalle por vigencia, y el total se
    obtiene sumando, que es lo que permite mostrar el acumulado sin volver a
    depender del Excel—.
    """
    salida = {v: {} for v in VIGENCIAS}
    for h in headers:
        n = _norm(h)
        if not n or "total" in n:
            continue
        for vig in VIGENCIAS:
            if str(vig) not in n:
                continue
            for campo, marcas in CAMPOS_PLATA.items():
                if any(mk in n for mk in marcas):
                    salida[vig].setdefault(campo, h)
    return salida


class Command(BaseCommand):
    help = ("Importa la Matriz de seguimiento PDL de la ALK (xlsx) y completa "
            "proyecto/metas/KPI faltantes. Seco por defecto.")

    def add_arguments(self, parser):
        parser.add_argument("xlsx_path", help="Ruta al archivo .xlsx de la matriz")
        parser.add_argument("--write", action="store_true",
                             help="Escribe de verdad. Sin esto, solo reporta.")
        parser.add_argument("--usuario", default=None,
                             help="Username de quien corre el import. Obligatorio con --write.")

    def handle(self, *args, **opts):
        try:
            import openpyxl
        except ImportError as e:
            raise CommandError("Falta openpyxl en el contenedor.") from e

        escribir = opts["write"]
        autor = None
        if escribir:
            from django.contrib.auth import get_user_model
            username = opts.get("usuario")
            if not username:
                raise CommandError(
                    "--write exige --usuario: la carga de un plan oficial sin "
                    "autor no queda defendible.")
            autor = get_user_model().objects.filter(username=username).first()
            if autor is None:
                raise CommandError(f"No existe el usuario «{username}».")

        wb = openpyxl.load_workbook(opts["xlsx_path"], data_only=True)
        prog_ws = wb["Programacion PDL 2025 - 2028"]
        seg_ws = wb["Seguimiento"]

        def filas(ws):
            headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
            for r in range(2, ws.max_row + 1):
                row = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
                if all(v is None for v in row):
                    continue
                yield dict(zip(headers, row))

        prog_rows = list(filas(prog_ws))
        seg_by_concat = {}
        for r in filas(seg_ws):
            concat = r.get("Codigo cocatenado")
            if concat is not None:
                seg_by_concat[str(concat)] = r

        self.stdout.write(self.style.MIGRATE_HEADING(
            f"=== Matriz PDL ALK: {len(prog_rows)} indicadores · "
            f"{'ESCRITURA' if escribir else 'SECO (sin --write no se persiste)'} ==="))

        from django.db import connection, transaction

        # ── Paso 2: completar metas existentes (por codigo_meta) ──
        with connection.cursor() as cur:
            cur.execute("SELECT codigo, codigo_meta FROM metas WHERE codigo_meta IS NOT NULL")
            meta_por_codigo_meta = {row[1]: row[0] for row in cur.fetchall()}

        backfill_n = 0
        divergencias = []
        creaciones_proyecto = []
        creaciones_indicador = []

        with connection.cursor() as cur:
            for r in prog_rows:
                proy_cod = int(r["Cód. Proyecto de Inversión SEGPLAN"])
                no_ind = str(int(r["No. Indicador "]))
                seg = None
                for concat, s in seg_by_concat.items():
                    if str(s.get("N° Proyecto de inversión")) == str(proy_cod) and \
                       str(s.get("Codigo indicador")) == no_ind:
                        seg = s
                        break
                codigo_meta_excel = seg.get("Codigo cocatenado") if seg else None
                meta_codigo = meta_por_codigo_meta.get(str(codigo_meta_excel)) if codigo_meta_excel else None

                sector = r.get("Sector")
                linea = r.get("Línea de Inversión ")
                concepto = r.get("Concepto de Gasto ")
                componente = r.get("COMPONENTE PROYECTO")
                anualizacion = r.get("Tipo de anualización meta")
                codprog = _leading_int(seg.get("Programa")) if seg else None
                nomprog = seg.get("Programa") if seg else None

                if meta_codigo is not None:
                    if not escribir:
                        # Preview de solo lectura: ¿le falta algo a esta meta?
                        cur.execute("""SELECT 1 FROM metas WHERE codigo=%s AND
                                       (sector IS NULL OR linea IS NULL OR concepto IS NULL
                                        OR componente IS NULL OR codind IS NULL)""",
                                    [meta_codigo])
                        if cur.fetchone():
                            backfill_n += 1
                        continue
                    # Backfill SOLO columnas NULL, nunca pisa lo que ya está escrito.
                    cur.execute("""
                        UPDATE metas SET
                            sector = COALESCE(sector, %s),
                            linea = COALESCE(linea, %s),
                            concepto = COALESCE(concepto, %s),
                            componente = COALESCE(componente, %s),
                            anualizacion = COALESCE(anualizacion, %s),
                            codind = COALESCE(codind, %s),
                            nomind = COALESCE(nomind, %s),
                            codprog = COALESCE(codprog, %s),
                            nomprog = COALESCE(nomprog, %s),
                            codproy = COALESCE(codproy, %s),
                            proyecto_codigo = COALESCE(proyecto_codigo, %s)
                        WHERE codigo = %s
                          AND (sector IS NULL OR linea IS NULL OR concepto IS NULL
                               OR componente IS NULL OR codind IS NULL)
                        RETURNING codigo
                    """, [sector, linea, concepto, componente, anualizacion,
                          int(no_ind), r.get("Indicador de producto"), codprog, nomprog,
                          proy_cod, proy_cod, meta_codigo])
                    if cur.fetchone():
                        backfill_n += 1
                        registrar_cambio_seguro(
                            autor, "meta", meta_codigo, "sector_linea_concepto_componente",
                            None, f"{sector} | {linea} | {concepto} | {componente}",
                            observacion="Backfill de metadatos SEGPLAN desde Matriz PDL ALK "
                                        f"(indicador {no_ind}, proyecto {proy_cod}).")

                    # Reporta divergencia de magnitud vigencia actual, sin tocarla.
                    cur.execute("""
                        SELECT k.id, k.meta_magnitud FROM presu_indicador_meta_proyecto k
                        JOIN meta_proyecto mp ON mp.id = k.meta_proyecto_id
                        WHERE mp.meta_id = %s AND k.activo
                    """, [meta_codigo])
                    for kpi_id, kpi_mag in cur.fetchall():
                        excel_2026 = _num(r.get("Magnitud Meta Reprogramacion 2026"))
                        if excel_2026 is not None and kpi_mag is not None and \
                           abs(float(kpi_mag) - excel_2026) > 0.01:
                            divergencias.append((proy_cod, no_ind, kpi_id, float(kpi_mag), excel_2026))
                    continue

                # ── No hay meta interna enganchada por codigo_meta ──
                #
                # Ojo con las metas AGRUPADAS: una sola fila de `metas` puede
                # cubrir dos indicadores SEGPLAN a la vez (medido en la meta 8
                # de 2377: "Impactar 1400 jóvenes" agrupa los indicadores 51 y
                # 52 con dos KPIs propios), y esa fila NUNCA tiene codigo_meta
                # —no puede, cubre dos códigos—. Sin este chequeo, el importador
                # las trata como "no enganchadas" y crea un KPI duplicado del
                # que ya existe (pasó la primera vez que corrió este comando,
                # con los indicadores 51/52 de 2377: se detectó y se revirtió
                # a mano). Si el proyecto YA tiene un KPI activo cuya magnitud
                # coincide con la de este indicador y cuya meta no tiene
                # codigo_meta, no se crea nada: se reporta para que una persona
                # decida si es el mismo o uno genuinamente nuevo.
                cur.execute("""
                    SELECT k.id, k.meta_magnitud FROM presu_indicador_meta_proyecto k
                    JOIN meta_proyecto mp ON mp.id = k.meta_proyecto_id
                    JOIN metas m ON m.codigo = mp.meta_id
                    WHERE mp.proyecto_id = (SELECT id FROM proyecto WHERE codigo = %s
                                             OR codigo = %s LIMIT 1)
                      AND k.activo AND m.codigo_meta IS NULL
                """, [str(proy_cod), str(proy_cod).zfill(7)])
                candidatas_sin_codigo = cur.fetchall()
                valores_excel = {_num(r.get(f"Magnitud Meta Reprogramacion{' POAI' if a == 2025 else ''} {a}"))
                                  for a in (2025, 2026, 2027, 2028)}
                valores_excel.add(_num(r.get("Meta  2025-2028")))
                posible_duplicado = next(
                    (kpi_id for kpi_id, mag in candidatas_sin_codigo
                     if mag is not None and any(v is not None and abs(float(mag) - v) < 0.01
                                                 for v in valores_excel)),
                    None)
                if posible_duplicado is not None:
                    self.stdout.write(self.style.WARNING(
                        f"    ⚠ proy {proy_cod} ind {no_ind}: KPI#{posible_duplicado} ya activo "
                        "sin codigo_meta y con magnitud igual — NO se crea, revisar a mano si es "
                        "el mismo indicador."))
                    continue

                # ── Genuinamente falta: crear la cadena ──
                creaciones_indicador.append((proy_cod, no_ind, r, seg))

        # Agrupar creaciones por proyecto para saber cuáles proyectos son nuevos.
        with connection.cursor() as cur:
            cur.execute("SELECT codigo FROM proyecto")
            proyectos_existentes = set()
            for (cod,) in cur.fetchall():
                try:
                    proyectos_existentes.add(int(cod))
                except (TypeError, ValueError):
                    pass

        proyectos_nuevos_necesarios = sorted({p for p, _, _, _ in creaciones_indicador} - proyectos_existentes)
        for cod in proyectos_nuevos_necesarios:
            sector = next(x[2]["Sector"] for x in creaciones_indicador if x[0] == cod)
            subgrupo_id, dudoso = _sector_a_subgrupo(sector, cod)
            creaciones_proyecto.append((cod, sector, subgrupo_id, dudoso))

        # ── Paso 2b: la PLATA por meta y vigencia ──
        #
        # Va a `presu_presupuesto_meta_vigencia` (DDL 020) y NO a
        # `sdp_meta_oficial`: el UNIQUE de aquella no incluye `fuente`, así que
        # escribir ahí no agrega una fuente en paralelo sino que PISA la fila
        # oficial. Ya pasó una vez y rompió 10 tests de apps.dashboard.
        #
        # Se guardan las cuatro columnas —proyectado PDL, apropiación POAI,
        # comprometido y girado— porque la cadena real de ejecución arranca en
        # la APROPIACIÓN, no en el proyectado (que es la meta aspiracional del
        # cuatrienio). Guardar solo la apropiación dejaría al cockpit sin poder
        # explicar la diferencia contra lo que mostraba antes.
        seg_headers = [seg_ws.cell(row=1, column=c).value
                       for c in range(1, seg_ws.max_column + 1)]
        cols_plata = _columnas_plata(seg_headers)
        archivo = opts["xlsx_path"].rsplit("/", 1)[-1]

        filas_plata = []
        for concat, srow in seg_by_concat.items():
            proy_cod = _leading_int(srow.get("N° Proyecto de inversión"))
            for vig in VIGENCIAS:
                valores = {campo: _num(srow.get(hdr))
                           for campo, hdr in cols_plata[vig].items()}
                if not any(v is not None for v in valores.values()):
                    continue
                filas_plata.append((str(concat), proy_cod, vig, valores))

        con_apropiacion = sum(1 for *_, v in filas_plata
                              if v.get("apropiacion_poai") is not None)
        self.stdout.write(self.style.MIGRATE_HEADING(
            f"\n[plata] {len(filas_plata)} filas meta×vigencia "
            f"({con_apropiacion} con apropiación POAI)"))

        if escribir and filas_plata:
            with connection.cursor() as cur:
                for concat, proy_cod, vig, valores in filas_plata:
                    cur.execute("""
                        INSERT INTO presu_presupuesto_meta_vigencia (
                            codigo_meta, proyecto_codigo, vigencia,
                            proyectado_pdl, apropiacion_poai, comprometido, girado,
                            fuente, archivo_origen, cargado_por_id
                        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        ON CONFLICT (codigo_meta, vigencia, fuente) DO UPDATE SET
                            proyecto_codigo  = EXCLUDED.proyecto_codigo,
                            proyectado_pdl   = EXCLUDED.proyectado_pdl,
                            apropiacion_poai = EXCLUDED.apropiacion_poai,
                            comprometido     = EXCLUDED.comprometido,
                            girado           = EXCLUDED.girado,
                            archivo_origen   = EXCLUDED.archivo_origen,
                            cargado_por_id   = EXCLUDED.cargado_por_id,
                            updated_at       = now()
                    """, [concat, proy_cod, vig,
                          valores.get("proyectado_pdl"),
                          valores.get("apropiacion_poai"),
                          valores.get("comprometido"),
                          valores.get("girado"),
                          "matriz_pdl_alk", archivo,
                          getattr(autor, "id", None)])
            # UNA sola entrada de auditoría para todo el bloque: 280 filas de
            # plata son un cargue, no 280 decisiones. La trazabilidad fina vive
            # en la tabla misma (archivo_origen + cargado_por_id + updated_at).
            registrar_cambio_seguro(
                autor, "presu_presupuesto_meta_vigencia", 0, "cargue",
                None, f"{len(filas_plata)} filas",
                observacion=(f"Presupuesto por meta y vigencia desde «{archivo}» "
                             f"({con_apropiacion} con apropiación POAI). "
                             "Fuente matriz_pdl_alk."))

        self.stdout.write(self.style.MIGRATE_HEADING(
            f"\n[metas] backfill de metadatos SEGPLAN: {backfill_n}"))
        if divergencias:
            self.stdout.write(self.style.WARNING(
                f"\n[divergencias] {len(divergencias)} KPI activos con magnitud "
                "distinta a la reprogramación 2026 de la ALK (NO se tocan solas):"))
            for proy, ind, kpi_id, actual, excel in divergencias:
                self.stdout.write(f"    proy {proy} ind {ind}: KPI#{kpi_id} tiene {actual}, "
                                  f"Excel 2026 dice {excel}")

        self.stdout.write(self.style.MIGRATE_HEADING(
            f"\n[proyectos nuevos] {len(creaciones_proyecto)}"))
        for cod, sector, subgrupo_id, dudoso in creaciones_proyecto:
            marca = " ⚠ revisar subgrupo" if dudoso else ""
            self.stdout.write(f"    {cod} ({sector}) → subgrupo_id={subgrupo_id}{marca}")

        self.stdout.write(self.style.MIGRATE_HEADING(
            f"\n[indicadores nuevos a crear] {len(creaciones_indicador)}"))

        if not escribir:
            self.stdout.write(self.style.WARNING(
                "\nSECO: nada se escribió (fuera del espejo, si hubo --write parcial "
                "arriba). Corré con --write --usuario <username> para persistir."))
            return

        # ── Paso 3: crear proyecto / metas / meta_proyecto / KPI faltantes ──
        with transaction.atomic(), connection.cursor() as cur:
            nombre_por_cod = {}
            for cod, sector, subgrupo_id, dudoso in creaciones_proyecto:
                nombre = next(x[2]["Nombre del Proyecto"] for x in creaciones_indicador if x[0] == cod)
                nombre_por_cod[cod] = nombre
                cur.execute("""
                    INSERT INTO proyecto (codigo, nombre, subgrupo_id)
                    VALUES (%s, %s, %s)
                    RETURNING id
                """, [str(cod), nombre, subgrupo_id])
                proy_id = cur.fetchone()[0]
                registrar_cambio_seguro(
                    autor, "proyecto", proy_id, "creacion", None, nombre,
                    proyecto_id=proy_id, subgrupo_id=subgrupo_id,
                    observacion=f"Creado desde Matriz PDL ALK (sector «{sector}»)."
                                + (" Subgrupo asignado por mejor coincidencia de sector: "
                                   "revisar con Alex." if dudoso else ""))

            cur.execute("SELECT id, codigo FROM proyecto")
            proyecto_id_por_cod = {}
            for pid, cod in cur.fetchall():
                try:
                    proyecto_id_por_cod[int(cod)] = pid
                except (TypeError, ValueError):
                    pass

            creados_metas = 0
            for proy_cod, no_ind, r, seg in creaciones_indicador:
                proyecto_id = proyecto_id_por_cod.get(proy_cod)
                if proyecto_id is None:
                    continue
                codigo_meta_excel = seg.get("Codigo cocatenado") if seg else None
                nomprog = seg.get("Programa") if seg else None
                cur.execute("""
                    INSERT INTO metas (
                        nombre, sector, linea, concepto, componente,
                        anualizacion, codind, nomind, codprog, nomprog,
                        codproy, proyecto_codigo, codigo_meta
                    ) VALUES (
                        %s, %s, %s, %s, %s,
                        %s, %s, %s, %s, %s, %s, %s, %s
                    ) RETURNING codigo
                """, [r.get("Meta proyecto 2025-2028 (PDL)"), r.get("Sector"),
                      r.get("Línea de Inversión "), r.get("Concepto de Gasto "),
                      r.get("COMPONENTE PROYECTO"), r.get("Tipo de anualización meta"),
                      int(no_ind), r.get("Indicador de producto"),
                      _leading_int(nomprog), nomprog, proy_cod, proy_cod,
                      str(codigo_meta_excel) if codigo_meta_excel else None])
                meta_codigo = cur.fetchone()[0]

                cur.execute("""
                    INSERT INTO meta_proyecto (id, meta_id, proyecto_id, fecha_inicio, fecha_fin)
                    VALUES (nextval('meta_proyecto_id_seq'), %s, %s, '2025-01-01', '2028-12-31')
                    RETURNING id
                """, [meta_codigo, proyecto_id])
                mp_id = cur.fetchone()[0]

                mag_2026 = _num(r.get("Magnitud Meta Reprogramacion 2026"))
                tipo_agg = "SUMA" if (r.get("Tipo de anualización meta") or "").strip().lower() == "suma" else "ULTIMO"
                cur.execute("""
                    INSERT INTO presu_indicador_meta_proyecto (
                        id, meta_proyecto_id, nombre, descripcion, unidad_medida,
                        meta_magnitud, tipo_agregacion, activo, created_at, updated_at
                    ) VALUES (
                        nextval('presu_indicador_meta_proyecto_id_seq'), %s, %s, %s, %s,
                        %s, %s, TRUE, now(), now()
                    ) RETURNING id
                """, [mp_id, (r.get("Meta proyecto 2025-2028 (PDL)") or "")[:250],
                      r.get("Indicador de producto"), "Unidades", mag_2026, tipo_agg])
                kpi_id = cur.fetchone()[0]

                registrar_cambio_seguro(
                    autor, "presu_indicador_meta_proyecto", kpi_id, "creacion",
                    None, r.get("Indicador de producto"),
                    proyecto_id=proyecto_id,
                    observacion=f"Creado desde Matriz PDL ALK: indicador {no_ind} del "
                                f"proyecto {proy_cod}, magnitud vigencia 2026={mag_2026}.")
                creados_metas += 1

        self.stdout.write(self.style.SUCCESS(
            f"\nOK: {len(creaciones_proyecto)} proyectos creados, {creados_metas} "
            f"metas/KPI creados, {backfill_n} metas completadas, firmado por {opts['usuario']}."))


def registrar_cambio_seguro(autor, entidad, entidad_id, campo, valor_anterior,
                             valor_nuevo, proyecto_id=None, subgrupo_id=None,
                             observacion=None):
    from apps.presupuesto.models.auditoria import AuditoriaDato
    from apps.presupuesto.services.auditoria import registrar_cambio
    registrar_cambio(
        usuario=autor, entidad=entidad, entidad_id=entidad_id, campo=campo,
        valor_anterior=valor_anterior, valor_nuevo=valor_nuevo,
        proyecto_id=proyecto_id, subgrupo_id=subgrupo_id,
        fuente=AuditoriaDato.MANUAL, observacion=observacion,
    )
