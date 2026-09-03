"""Importa la «Alerta» de cumplimiento por meta desde la Matriz PDL (xlsx).

    docker exec innova_k python manage.py importar_alerta_metas_pdl archivo.xlsx
    docker exec innova_k python manage.py importar_alerta_metas_pdl archivo.xlsx --write --usuario <username>

SECO POR DEFECTO, firmado e IDEMPOTENTE — misma convención que
`importar_matriz_pdl_alk`, del que es hermano: mismo archivo de entrada,
hoja distinta.

POR QUÉ EXISTE. Pedido: un filtro de proyectos por estado de ejecución
(Crítico / En ejecución según cronograma / Ejecutada / Desierta / Sin
magnitud contratada). Esa taxonomía la calcula la ALK en la hoja «Alertas»
del mismo Excel que ya alimenta `importar_matriz_pdl_alk`, tabla de detalle
en A68:G146 (78 metas evaluadas), y hoy no está en ningún lado de la base.

QUÉ HACE. Lee la hoja «Alertas», encuentra la fila de encabezado («Proyecto
| Meta | Contratada | Ejecutada | Cumplimiento % | Diferencia | Alerta») y,
por cada fila de detalle, engancha la meta interna por DOS señales —nunca
una sola— dentro del mismo proyecto:

  1. Similaridad de texto entre la columna «Meta» y `metas.nombre` (que
     guarda la misma frase de origen, «Meta proyecto 2025-2028 (PDL)»,
     solo que la hoja Alertas la reformatea: "Dotar 74 sedes..." en
     Programación vs "Dotar 74 Sede(s)..." en Alertas — no son la misma
     cadena, pero sí la misma meta).
  2. La magnitud «Contratada» de Alertas contra `meta_magnitud` del KPI
     vivo de esa meta, como desempate cuando hay más de una meta candidata
     con similaridad de texto parecida.

Si ninguna meta del proyecto pasa el umbral de similaridad, la fila se
reporta como «sin enganche» y NO se escribe: no hay engaño posible en
adivinar la meta equivocada de un reporte que un concejal puede leer.

NO recalcula la alerta. El umbral (<50 % / 50-99,9 % / ≥100 %) que separa
Crítico / En ejecución / Ejecutada es de la ALK, calculado en su propia
hoja; acá se importa el VALOR ya resuelto, igual que el resto de esta
familia de importadores no reconstruye lo que la fuente ya entrega hecho.

Va a `presu_presupuesto_meta_vigencia` (DDL 021, columnas nuevas sobre la
MISMA tabla de DDL 020 — es el mismo cargue de la misma fuente, mismo
`codigo_meta` + `vigencia` + `fuente`) con `vigencia=2025` porque el
título de la hoja lo dice expresamente («Alertas de cumplimiento de metas
2025»).
"""
import difflib
import re

from django.core.management.base import BaseCommand, CommandError

VIGENCIA_ALERTA = 2025
FUENTE = "matriz_pdl_alk"

#: Único filtro real: que la palabra que llega del Excel sea una de las
#: cinco que la pantalla sabe pintar. No es un CHECK de base (ver DDL 021)
#: a propósito — acá SÍ conviene frenar, porque es la última oportunidad
#: de detectar una categoría nueva antes de que el frontend reciba un
#: valor que no sabe agrupar.
ALERTAS_VALIDAS = {
    # El valor real de la celda es «de acuerdo a», no «según» — se verificó
    # corriendo el importador en seco contra el Excel real: la primera
    # versión de esta lista, escrita a mano leyendo el título del gráfico
    # («EN EJECUCIÓN SEGÚN CRONOGRAMA»), no coincidía con el dato de la
    # columna G y el importador frenó él solo, como tiene que hacerlo.
    "Crítico", "En ejecución de acuerdo a cronograma", "Ejecutada",
    "Desierta", "Sin magnitud contratada",
}

#: Bajo este umbral de similaridad de texto, la fila se reporta y no se
#: escribe. Medido a mano contra las 78 filas de la hoja: la reformulación
#: "Dotar 74 sedes..." → "Dotar 74 Sede(s)..." da ~0.90; dos metas
#: distintas del mismo proyecto rara vez superan 0.45.
UMBRAL_SIMILARIDAD = 0.55

#: Piso de similaridad SOLO para el caso «única meta interna del proyecto».
#: Más bajo que `UMBRAL_SIMILARIDAD` porque la reformulación de la hoja
#: Alertas baja la similaridad de una meta genuina (medido: 0.36-0.46 en
#: varios casos correctos), pero no tan bajo como para aceptar cualquier
#: cosa — 0.25-0.26 (metas de becas de 2377 contra la única meta de
#: dotación que sí existe) es «tema distinto», no «texto reformulado».
PISO_UNICO = 0.30

#: Enganches leídos a mano, uno por uno, para los casos donde ni el texto ni
#: la magnitud alcanzan — nunca para ahorrarse revisar. Clave: número de fila
#: del Excel (columna A empieza en la fila del header + 1). Valor:
#: `codigo_meta` interno.
#:
#: `27112` = "IVC" en `metas.nombre`: acrónimo de Inspección, Vigilancia y
#: Control. La fila 127 dice literalmente «Realizar 4 estrategias de
#: inspección, vigilancia y control (una por vigencia)» — es la meta, pero
#: contra un acrónimo de 3 letras la similaridad de texto da 0.05 y no hay
#: magnitud que comparar (`meta_magnitud` de 27112 es NULL). Verificado a
#: mano contra las 3 metas del proyecto 2711: `27111` («Realizar 4
#: estrategias de fortalecimiento institucional») ya engancha con otra fila
#: por texto, y `27113` («Intervenir 1 sede administrativa») no tiene
#: relación temática ninguna — 27112 es la única que queda y la única que
#: encaja.
#:
#: `27902` = «Mantenimiento de parques de proximidad» (magnitud 13,
#: idéntica a la Contratada de la fila 109). El texto engaña acá: la fila
#: 109 dice «...con acciones de mejoramiento, MANTENIMIENTO y/o dotación» y
#: comparte la frase «parques de la red de proximidad» con `27901`
#: («Construir 1000 m2 de Parques...», magnitud 0), así que por similaridad
#: de texto puro 27901 gana (0.62 vs 0.40) — pero 27901 ya es la fila 94
#: («Construir 1000 m2...», con similaridad más alta todavía) y la
#: coincidencia exacta de magnitud (13=13) dice que la fila 109 es
#: mantenimiento, no construcción.
#:
#: `27452`/`27454` (proyecto 2745, «Kennedy Camina Hacia la Convivencia»):
#: dos metas de magnitud 1.0 cada una, sin desempate numérico posible.
#: - Fila 90 «Implementar 4 acciones pedagógicas para la gestión de
#:   conflictividades...» → `27452` «Implementar 1 acción pedagógica»: el
#:   verbo Y el sustantivo calzan exactos, contra `27454` que solo comparte
#:   el tema general.
#: - Fila 88 «Fortalecer 4 programas de abordaje de conflictividad escolar
#:   para la convivencia con enfoque restaurativo.» → `27454` «Fortalecer 1
#:   programa de conflictividad escolar»: el verbo Y el sustantivo calzan
#:   casi literal, contra `27451` «Ejecutar 1 programa comunitario con
#:   enfoque restaurativo» que solo comparte la frase final de contexto.
#:
#: `27062` (proyecto 2706): fila 124 «Intervenir 4 equipamientos de
#: seguridad y acceso a la justicia...» vs «Intervenir 4 equipamientos de
#: seguridad y justicia...» — 0.54, un punto por debajo del umbral por la
#: frase «acceso a la», sin ambigüedad real (la otra meta del proyecto,
#: `27061`, es de dotaciones y ya la toma la fila 125 por texto).
#:
#: `27841` (proyecto 2784, Banco de Iniciativas): fila 144 «Beneficiar 280
#: Colectivo(s) u organizaciones recreo deportivas inscritas en el Banco…»
#: vs «Beneficiar a 280 colectivos recreodeportivos en la localidad de
#: Kennedy» — 0.51, bajo el umbral por la reformulación, pero el número 280
#: y el tema (colectivos recreodeportivos del Banco) son inequívocos; la
#: Contratada es 0 (la fila es «Desierta»: nada se contrató), por eso la
#: magnitud no puede confirmar ni contradecir acá.
OVERRIDE_MANUAL: dict[int, str] = {
    127: "27112",
    109: "27902",
    90: "27452",
    88: "27454",
    124: "27062",
    144: "27841",
}


def _num(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _leading_int(s):
    m = re.match(r"\s*(\d+)", str(s or ""))
    return int(m.group(1)) if m else None


def _normaliza(s):
    s = (s or "").lower()
    s = re.sub(r"[^\w\s]", " ", s, flags=re.UNICODE)
    return " ".join(s.split())


def _similaridad(a, b):
    return difflib.SequenceMatcher(None, _normaliza(a), _normaliza(b)).ratio()


class Command(BaseCommand):
    help = ("Importa la «Alerta» de cumplimiento por meta desde la hoja "
            "«Alertas» de la Matriz PDL ALK (xlsx). Seco por defecto.")

    def add_arguments(self, parser):
        parser.add_argument("xlsx_path", help="Ruta al archivo .xlsx de la matriz")
        parser.add_argument("--write", action="store_true",
                             help="Escribe de verdad. Sin esto, solo reporta.")
        parser.add_argument("--usuario", default=None,
                             help="Username de quien corre el import. Obligatorio con --write.")
        parser.add_argument("--umbral", type=float, default=UMBRAL_SIMILARIDAD,
                             help=f"Umbral de similaridad de texto (default {UMBRAL_SIMILARIDAD}).")

    def handle(self, *args, **opts):
        try:
            import openpyxl
        except ImportError as e:
            raise CommandError("Falta openpyxl en el contenedor.") from e

        escribir = opts["write"]
        umbral = opts["umbral"]
        autor = None
        if escribir:
            from django.contrib.auth import get_user_model
            username = opts.get("usuario")
            if not username:
                raise CommandError(
                    "--write exige --usuario: la carga de un plan oficial sin "
                    "autor no queda defendible.")
            autor = get_user_model().objects.filter(username=username).first()
            if autor is None:
                raise CommandError(f"No existe el usuario «{username}».")

        wb = openpyxl.load_workbook(opts["xlsx_path"], data_only=True)
        if "Alertas" not in wb.sheetnames:
            raise CommandError(
                f"El archivo no tiene hoja «Alertas» (hojas: {wb.sheetnames}).")
        ws = wb["Alertas"]

        # ── Encuentra la fila de encabezado del detalle: es la única fila
        # de toda la hoja cuya columna A dice literalmente «Proyecto» y cuya
        # columna G dice «Alerta». El resto de la hoja son gráficos y
        # tarjetas resumen que no tienen esa forma. ──
        header_row = None
        for r in range(1, ws.max_row + 1):
            if (ws.cell(row=r, column=1).value == "Proyecto"
                    and ws.cell(row=r, column=7).value == "Alerta"):
                header_row = r
                break
        if header_row is None:
            raise CommandError(
                "No se encontró la tabla de detalle (fila con «Proyecto»…«Alerta») "
                "en la hoja «Alertas». ¿Cambió el layout del Excel?")

        filas_excel = []
        for r in range(header_row + 1, ws.max_row + 1):
            proyecto_txt = ws.cell(row=r, column=1).value
            if proyecto_txt is None:
                continue
            filas_excel.append({
                "fila": r,
                "proyecto_txt": str(proyecto_txt),
                "meta_txt": ws.cell(row=r, column=2).value,
                "contratada": _num(ws.cell(row=r, column=3).value),
                "ejecutada": _num(ws.cell(row=r, column=4).value),
                "cumplimiento_pct": _num(ws.cell(row=r, column=5).value),
                "alerta": ws.cell(row=r, column=7).value,
            })

        self.stdout.write(self.style.MIGRATE_HEADING(
            f"=== Alertas por meta: {len(filas_excel)} filas de detalle "
            f"(fila {header_row + 1} a {ws.max_row}) · "
            f"{'ESCRITURA' if escribir else 'SECO (sin --write no se persiste)'} ==="))

        # ── Valida las categorías ANTES de tocar nada ──
        desconocidas = {f["alerta"] for f in filas_excel} - ALERTAS_VALIDAS - {None}
        if desconocidas:
            raise CommandError(
                f"La hoja trae categorías de Alerta que la pantalla no conoce: "
                f"{sorted(desconocidas)}. Revisar ALERTAS_VALIDAS antes de importar.")

        from django.db import connection

        # ── Candidatas por proyecto: codigo_meta, nombre (texto de origen) y
        # meta_magnitud del KPI vivo, para el enganche por similaridad + magnitud. ──
        with connection.cursor() as cur:
            cur.execute("""
                SELECT m.codigo_meta, m.proyecto_codigo, m.nombre, k.meta_magnitud
                FROM metas m
                JOIN meta_proyecto mp ON mp.meta_id = m.codigo
                LEFT JOIN presu_indicador_meta_proyecto k
                       ON k.meta_proyecto_id = mp.id AND k.activo
                WHERE m.proyecto_codigo IS NOT NULL
            """)
            candidatas_por_proyecto: dict[int, list[dict]] = {}
            for codigo_meta, proy_cod, nombre, magnitud in cur.fetchall():
                if codigo_meta is None or proy_cod is None:
                    continue
                candidatas_por_proyecto.setdefault(int(proy_cod), []).append({
                    "codigo_meta": codigo_meta, "nombre": nombre, "magnitud": magnitud,
                })

        # ── Enganche: asignación bipartita VORAZ, no «mejor candidata por
        # fila» a secas. ──
        #
        # La primera versión de este importador eligió la mejor candidata
        # POR FILA sin mirar qué elegían las demás filas del mismo proyecto,
        # y en la corrida real dos pares de filas del Excel engancharon con
        # LA MISMA meta interna (proyectos 2790 y 2711): la segunda fila
        # pisó en silencio el `alerta` que acababa de escribir la primera,
        # vía el mismo UPSERT que blinda `codigo_meta` — que es justo lo que
        # ese UPSERT no debía dejar pasar. Se detectó DESPUÉS de escribir,
        # comparando cuántas filas se enganchaban contra cuántas filas
        # quedaron con `alerta IS NOT NULL` (67 vs 65) — la única señal de
        # que algo se pisó fue el número, no un error.
        #
        # Con la asignación voraz global: se ordenan TODAS las parejas
        # (fila, meta candidata) de todo el archivo por similaridad
        # descendente, y se van tomando de mayor a menor SOLO si ni la fila
        # ni la meta ya están usadas. Una meta interna nunca recibe dos
        # filas del Excel; la fila que pierde la carrera por «su» mejor
        # candidata queda sin enganche y se reporta, en vez de conformarse
        # con la segunda mejor y arriesgar otro enganche torcido.
        pares: list[tuple[float, dict, str, int]] = []
        sin_candidatas: list[tuple[dict, str]] = []
        ambiguas: list[tuple[dict, str]] = []
        for f in filas_excel:
            proy_cod = _leading_int(f["proyecto_txt"])
            candidatas = candidatas_por_proyecto.get(proy_cod, [])
            if not candidatas:
                sin_candidatas.append((f, "el proyecto no tiene metas internas cargadas"))
                continue

            # ── Override manual, documentado — no un ajuste del algoritmo ──
            #
            # `27112` se llama "IVC" en `metas.nombre` — es un acrónimo de
            # "Inspección, Vigilancia y Control", que es justo el texto de
            # la fila 127. La similaridad de texto contra un acrónimo de 3
            # letras da 0.05: ningún umbral razonable lo salva sin also
            # aceptar basura. No es el algoritmo adivinando — es una
            # correspondencia leída a mano, igual que
            # `PROYECTO_SUBGRUPO_OVERRIDE` en `importar_matriz_pdl_alk`.
            override = OVERRIDE_MANUAL.get(f["fila"])
            if override is not None:
                pares.append((99.0, f, override, proy_cod))
                continue

            puntajes = []
            for c in candidatas:
                sim = _similaridad(f["meta_txt"], c["nombre"])
                puntajes.append([sim, c, False])
            puntajes.sort(key=lambda t: t[0], reverse=True)

            # ── Ancla de magnitud: RESCATE, no árbitro. ──
            #
            # Solo entra a jugar cuando el texto SOLO no alcanza el umbral
            # para NINGÚN candidato — nunca para desempatar contra un texto
            # que ya viene ganando por su cuenta. La primera versión de este
            # rescate le daba el mismo peso a la magnitud siempre, y en los
            # proyectos 2780/2794 dos filas del Excel compartían el mismo
            # número de «Contratada» por coincidencia (15 y 400): la ancla le
            # arrebató la meta correcta a una fila con ~0.95 de similaridad
            # de texto para dársela a otra con la magnitud pero el tema
            # equivocado. Medido, se detectó porque esa fila ganadora por
            # texto puro pasó a «perdió la carrera» — la señal de que algo
            # se invirtió al revés.
            texto_ya_alcanza = puntajes[0][0] >= umbral
            ancla = False
            if not texto_ya_alcanza and f["contratada"] is not None:
                exactas = [i for i, (_, c, _) in enumerate(puntajes)
                           if c["magnitud"] is not None
                           and abs(float(c["magnitud"]) - f["contratada"]) < 0.01]
                if len(exactas) == 1:
                    i = exactas[0]
                    puntajes[i][0] += 10.0
                    puntajes[i][2] = True
                    ancla = True
                    puntajes.sort(key=lambda t: t[0], reverse=True)

            # Único candidato del proyecto, CON un piso de similaridad —no
            # cualquier similaridad—: con 23773 como única meta interna de
            # 2377, las filas 118/119 («Beneficiar 700 Estudiante(s)…» de
            # BECAS) tienen 0.25 de similaridad contra 23773 («Dotar 74
            # sedes…», de DOTACIÓN) — es baja porque son temas DISTINTOS, no
            # porque falte texto en común. Eso no es «no hay entre quién
            # elegir mal», es que la meta de becas todavía no existe en la
            # base (ver docs/bitácora Jóvenes a la E, metas 23771/23772
            # pendientes) y forzar el enganche pisaría la meta de dotación
            # con datos de otra cosa. `PISO_UNICO` es más laxo que el umbral
            # general porque la reformulación de la hoja SÍ baja la
            # similaridad de una meta genuina (medido 0.36-0.46 en varios
            # casos correctos), pero 0.25-0.26 ya cruzó a «otro tema».
            unico = len(candidatas) == 1 and puntajes[0][0] >= PISO_UNICO

            mejor_sim, mejor, ancla_ganadora = puntajes[0]

            if not ancla_ganadora and not unico and mejor_sim < umbral:
                sin_candidatas.append((f, f"mejor candidata a {mejor_sim:.2f} de similaridad "
                                           f"(umbral {umbral}), por debajo del corte"))
                continue

            # Ambigüedad DENTRO del proyecto: dos metas casi igual de
            # parecidas para esta misma fila. Ninguna asignación global
            # puede resolver esto — se excluye de entrada. No aplica con
            # ancla: esa evidencia no admite empate posible (es la única
            # exacta, por construcción, y ya se descartó si el texto solo
            # tenía un candidato por encima del umbral).
            if not ancla_ganadora and len(puntajes) > 1 and abs(puntajes[1][0] - mejor_sim) < 0.03:
                ambiguas.append((f, f"empate entre 2+ metas del proyecto "
                                     f"({mejor_sim:.2f} vs {puntajes[1][0]:.2f}) — revisar a mano"))
                continue

            pares.append((mejor_sim, f, mejor["codigo_meta"], proy_cod))

        pares.sort(key=lambda t: t[0], reverse=True)
        filas_usadas: set[int] = set()
        metas_usadas: set[str] = set()
        enganchadas = []
        perdio_la_carrera: list[tuple[dict, str]] = []
        for sim, f, codigo_meta, proy_cod in pares:
            if f["fila"] in filas_usadas:
                continue
            if codigo_meta in metas_usadas:
                perdio_la_carrera.append(
                    (f, f"la meta {codigo_meta} ya se le asignó a otra fila del Excel "
                        f"con más similaridad — revisar a mano si son dos metas distintas"))
                continue
            filas_usadas.add(f["fila"])
            metas_usadas.add(codigo_meta)
            enganchadas.append((f, codigo_meta, proy_cod, sim))

        sin_enganche = sin_candidatas + ambiguas + perdio_la_carrera
        # Reordena por número de fila para que el reporte se lea en el mismo
        # orden que el Excel, sin importar en qué balde cayó cada una.
        sin_enganche.sort(key=lambda t: t[0]["fila"])

        self.stdout.write(self.style.MIGRATE_HEADING(
            f"\n[enganche] {len(enganchadas)} de {len(filas_excel)} filas engancharon "
            f"con una meta interna (umbral {umbral})"))
        if sin_enganche:
            self.stdout.write(self.style.WARNING(
                f"\n[sin enganche] {len(sin_enganche)} filas NO se importan:"))
            for f, motivo in sin_enganche:
                self.stdout.write(
                    f"    fila {f['fila']} · {f['proyecto_txt']} · "
                    f"«{(f['meta_txt'] or '')[:70]}» → {motivo}")

        por_alerta = {}
        for f, *_r in enganchadas:
            por_alerta[f["alerta"]] = por_alerta.get(f["alerta"], 0) + 1
        self.stdout.write(self.style.MIGRATE_HEADING("\n[por alerta, de lo enganchado]"))
        for alerta, n in sorted(por_alerta.items()):
            self.stdout.write(f"    {alerta}: {n}")

        if not escribir:
            self.stdout.write(self.style.WARNING(
                "\nSECO: nada se escribió. Corré con --write --usuario <username> "
                "para persistir lo enganchado (lo sin enganche nunca se escribe solo "
                "con más --write: hay que resolver el enganche a mano o subir el "
                "umbral con --umbral una vez revisado)."))
            return

        archivo = opts["xlsx_path"].rsplit("/", 1)[-1]
        with connection.cursor() as cur:
            for f, codigo_meta, proy_cod, _sim in enganchadas:
                cur.execute("""
                    INSERT INTO presu_presupuesto_meta_vigencia (
                        codigo_meta, proyecto_codigo, vigencia,
                        alerta, magnitud_contratada, magnitud_ejecutada, cumplimiento_pct,
                        fuente, archivo_origen, cargado_por_id
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (codigo_meta, vigencia, fuente) DO UPDATE SET
                        proyecto_codigo      = EXCLUDED.proyecto_codigo,
                        alerta               = EXCLUDED.alerta,
                        magnitud_contratada  = EXCLUDED.magnitud_contratada,
                        magnitud_ejecutada   = EXCLUDED.magnitud_ejecutada,
                        cumplimiento_pct     = EXCLUDED.cumplimiento_pct,
                        archivo_origen       = EXCLUDED.archivo_origen,
                        cargado_por_id       = EXCLUDED.cargado_por_id,
                        updated_at           = now()
                """, [codigo_meta, proy_cod, VIGENCIA_ALERTA,
                      f["alerta"], f["contratada"], f["ejecutada"], f["cumplimiento_pct"],
                      FUENTE, archivo, getattr(autor, "id", None)])

        registrar_cambio_seguro(
            autor, "presu_presupuesto_meta_vigencia", 0, "alerta_cargue",
            None, f"{len(enganchadas)} filas",
            observacion=(f"Alerta de cumplimiento por meta desde «{archivo}» "
                         f"({len(enganchadas)} de {len(filas_excel)} filas enganchadas, "
                         f"vigencia {VIGENCIA_ALERTA}). Fuente {FUENTE}."))

        self.stdout.write(self.style.SUCCESS(
            f"\nOK: {len(enganchadas)} filas escritas, {len(sin_enganche)} sin enganche "
            f"(revisar a mano), firmado por {opts['usuario']}."))


def registrar_cambio_seguro(autor, entidad, entidad_id, campo, valor_anterior,
                             valor_nuevo, observacion=None):
    from apps.presupuesto.models.auditoria import AuditoriaDato
    from apps.presupuesto.services.auditoria import registrar_cambio
    registrar_cambio(
        usuario=autor, entidad=entidad, entidad_id=entidad_id, campo=campo,
        valor_anterior=valor_anterior, valor_nuevo=valor_nuevo,
        fuente=AuditoriaDato.MANUAL, observacion=observacion,
    )
