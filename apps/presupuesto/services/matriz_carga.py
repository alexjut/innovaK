"""El motor de la CARGA de la Matriz PDL (pieza 4).

Convierte `activo` / `carga_origen_id` / `carga_retiro_id` —hoy tres columnas
que nadie escribe— en la regla del plan:

    **La carga nunca borra.** Lo que desaparece de la matriz se marca
    `activo = FALSE` apuntando a la carga que lo retiró.

Dos operaciones, en este orden y nunca fusionadas:

    previsualizar(ruta, corte, usuario)  → una carga en `borrador` con su diff
    aplicar(carga_id, usuario)           → la escribe, en UNA transacción

Están separadas a propósito: el diff tiene que poder mirarse y decidirse. Si
`previsualizar` aplicara, el estado `borrador` no significaría nada y las tres
pantallas de la Fase C no tendrían dónde apoyarse.

QUÉ CUBRE Y QUÉ NO
------------------
Cubre la JERARQUÍA —sectores, objetivos, programas—, que es donde vive la regla
de retiro: son entidades con nombre propio que pueden dejar de existir. NO
toca la plata ni las magnitudes: ésas siguen entrando por
`importar_matriz_pdl_alk`, que ya es idempotente y no pisa lo escrito. Fusionar
las dos cosas en un solo paso mezclaría «cambió el catálogo del Plan» con
«llegaron cifras nuevas», que se revisan distinto y se equivocan distinto.

LAS FILAS SEMBRADAS ANTES DE ESTO
---------------------------------
Los 13 sectores, 5 objetivos y 22 programas que ya están tienen
`carga_origen_id = NULL`: se sembraron con los DDL 023 y 024, antes de que
existiera el registro de cargas. Se dejan así. Ponerles la primera carga que
pase diría que esa carga los trajo, y es falso — un NULL que significa «venía
de antes» es más honesto que una atribución inventada.
"""
import unicodedata
from datetime import datetime, timezone

from django.db import transaction

from apps.presupuesto.models import (MatrizPDLCarga, ObjetivoEstrategico,
                                     ProgramaPDL, Sector)

HOJA_PROG = "Programacion PDL 2025 - 2028"
HOJA_SEG = "Seguimiento"

COL_SECTOR = 2       # Programacion · C
COL_OBJETIVO = 0     # Seguimiento  · A
COL_PROGRAMA = 1     # Seguimiento  · B

ENCABEZADOS = {
    HOJA_PROG: {COL_SECTOR: "Sector"},
    HOJA_SEG: {COL_OBJETIVO: "Objetivo Estrategico", COL_PROGRAMA: "Programa"},
}


class CargaError(Exception):
    """Algo impide seguir. El mensaje va tal cual a la pantalla."""


def norm_texto(v):
    """Mayúsculas, sin tildes, sin espacios dobles. UNA implementación."""
    if v is None:
        return ""
    s = unicodedata.normalize("NFKD", str(v))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return " ".join(s.upper().split())


def _partir(texto):
    """«3 - Bogotá confía…» → (3, 'Bogotá confía…')."""
    limpio = " ".join(str(texto).split())
    cabeza, sep, resto = limpio.partition("-")
    if not sep or not cabeza.strip().isdigit():
        raise CargaError(
            f"«{limpio}» no tiene la forma «N - nombre». La matriz cambió de "
            f"convención: revisar docs/operacion/matriz_pdl_mapeo.md.")
    return int(cabeza.strip()), resto.strip()


def leer_jerarquia(ruta):
    """Lee las tres entidades de la matriz. No toca la base.

    Valida los encabezados por nombre antes de leer: si la ALK mueve una
    columna, esto aborta en vez de cargar la columna de al lado en silencio.
    """
    try:
        import openpyxl
    except ImportError:  # pragma: no cover - depende del entorno
        raise CargaError("Falta openpyxl para leer el Excel.")

    try:
        wb = openpyxl.load_workbook(ruta, data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001 - el motivo va a pantalla
        raise CargaError(f"No se pudo abrir el archivo: {exc}")

    for hoja in (HOJA_PROG, HOJA_SEG):
        if hoja not in wb.sheetnames:
            raise CargaError(
                f"Falta la hoja «{hoja}». El archivo trae: {wb.sheetnames}")

    def filas(hoja):
        it = wb[hoja].iter_rows(values_only=True)
        cabecera = next(it, ()) or ()
        for idx, esperado in ENCABEZADOS[hoja].items():
            visto = cabecera[idx] if idx < len(cabecera) else None
            if norm_texto(visto) != norm_texto(esperado):
                raise CargaError(
                    f"En «{hoja}» la columna {idx} debería ser «{esperado}» y "
                    f"dice «{visto}».")
        for fila in it:
            if fila and any(c is not None for c in fila):
                yield fila

    sectores, objetivos, programas = {}, {}, {}
    for fila in filas(HOJA_PROG):
        crudo = fila[COL_SECTOR]
        if crudo and str(crudo).strip():
            nombre = " ".join(str(crudo).split())
            sectores[norm_texto(nombre)] = nombre

    for fila in filas(HOJA_SEG):
        if not fila[COL_OBJETIVO] or not fila[COL_PROGRAMA]:
            continue
        cod_o, nom_o = _partir(fila[COL_OBJETIVO])
        cod_p, nom_p = _partir(fila[COL_PROGRAMA])
        objetivos[cod_o] = nom_o
        previo = programas.get(cod_p)
        if previo and previo[1] != cod_o:
            raise CargaError(
                f"El programa {cod_p} aparece bajo los objetivos {previo[1]} y "
                f"{cod_o}. El modelo asume UN padre por programa.")
        programas[cod_p] = (nom_p, cod_o)

    if not sectores or not objetivos or not programas:
        raise CargaError(
            "El archivo no trae jerarquía: "
            f"{len(sectores)} sectores, {len(objetivos)} objetivos, "
            f"{len(programas)} programas. No se registra una carga vacía.")
    return {"sectores": sectores, "objetivos": objetivos, "programas": programas}


def _diff_sectores(leidos):
    vivos = {norm_texto(s.nombre_oficial): s for s in Sector.objects.all()}
    altas = [n for k, n in leidos.items() if k not in vivos]
    reactivar = [vivos[k].nombre_oficial for k in leidos if k in vivos and not vivos[k].activo]
    retiros = [s.nombre_oficial for k, s in vivos.items() if k not in leidos and s.activo]
    return {"altas": sorted(altas), "reactivaciones": sorted(reactivar),
            "cambios": [], "retiros": sorted(retiros)}


def _diff_objetivos(leidos):
    vivos = {o.codigo: o for o in ObjetivoEstrategico.objects.all()}
    altas = [{"codigo": c, "nombre": n} for c, n in leidos.items() if c not in vivos]
    cambios = [{"codigo": c, "de": vivos[c].nombre, "a": n}
               for c, n in leidos.items() if c in vivos and vivos[c].nombre != n]
    reactivar = [c for c in leidos if c in vivos and not vivos[c].activo]
    retiros = [{"codigo": c, "nombre": o.nombre}
               for c, o in vivos.items() if c not in leidos and o.activo]
    return {"altas": sorted(altas, key=lambda d: d["codigo"]),
            "cambios": sorted(cambios, key=lambda d: d["codigo"]),
            "reactivaciones": sorted(reactivar),
            "retiros": sorted(retiros, key=lambda d: d["codigo"])}


def _diff_programas(leidos):
    vivos = {p.codigo: p for p in ProgramaPDL.objects.select_related("objetivo")}
    altas, cambios = [], []
    for cod, (nombre, cod_obj) in leidos.items():
        actual = vivos.get(cod)
        if actual is None:
            altas.append({"codigo": cod, "nombre": nombre, "objetivo": cod_obj})
            continue
        campos = {}
        if actual.nombre != nombre:
            campos["nombre"] = {"de": actual.nombre, "a": nombre}
        if actual.objetivo.codigo != cod_obj:
            # Un programa que cambia de objetivo es reorganización del Plan, no
            # una corrección de texto: se marca aparte para que se vea.
            campos["objetivo"] = {"de": actual.objetivo.codigo, "a": cod_obj}
        if campos:
            cambios.append({"codigo": cod, "campos": campos})
    reactivar = [c for c in leidos if c in vivos and not vivos[c].activo]
    retiros = [{"codigo": c, "nombre": p.nombre}
               for c, p in vivos.items() if c not in leidos and p.activo]
    return {"altas": sorted(altas, key=lambda d: d["codigo"]),
            "cambios": sorted(cambios, key=lambda d: d["codigo"]),
            "reactivaciones": sorted(reactivar),
            "retiros": sorted(retiros, key=lambda d: d["codigo"])}


def calcular_diff(leidos):
    """El diff contra el estado vigente. Solo lee."""
    return {
        "sector": _diff_sectores(leidos["sectores"]),
        "objetivo": _diff_objetivos(leidos["objetivos"]),
        "programa": _diff_programas(leidos["programas"]),
    }


def _contar(diff):
    suma = lambda k: sum(len(bloque.get(k, [])) for bloque in diff.values())  # noqa: E731
    # Las reactivaciones cuentan como CAMBIO, no como alta: la fila ya existía
    # y conserva su `carga_origen`. Contarlas como altas diría que la carga las
    # trajo, y solo las devolvió a la vida.
    return {"n_altas": suma("altas"),
            "n_cambios": suma("cambios") + suma("reactivaciones"),
            "n_retiros": suma("retiros"),
            "n_errores": 0}


def previsualizar(ruta, corte_oficial, usuario_id=None, archivo_nombre=None):
    """Registra la carga en `borrador` con su diff. NO aplica nada."""
    h = MatrizPDLCarga.hash_de(ruta)
    previa = MatrizPDLCarga.objects.filter(hash_sha256=h).first()
    if previa:
        raise CargaError(
            f"Este archivo ya se subió: carga {previa.id} del "
            f"{previa.subido_at:%Y-%m-%d}, estado «{previa.estado}». "
            f"Si es un corte nuevo, el archivo tiene que ser distinto.")

    diff = calcular_diff(leer_jerarquia(ruta))
    import os
    return MatrizPDLCarga.objects.create(
        archivo_nombre=archivo_nombre or os.path.basename(ruta),
        hash_sha256=h,
        archivo_bytes=os.path.getsize(ruta),
        corte_oficial=corte_oficial,
        estado=MatrizPDLCarga.BORRADOR,
        diff=diff,
        subido_por_id=usuario_id,
        **_contar(diff),
    )


@transaction.atomic
def aplicar(carga_id, usuario_id=None):
    """Escribe el diff guardado. Todo o nada.

    Se aplica el diff que se GUARDÓ, no uno recalculado: lo que se aplica tiene
    que ser exactamente lo que alguien miró y aprobó. Si la base cambió entre
    la previsualización y esto, se ve en el resultado y se vuelve a subir.
    """
    carga = MatrizPDLCarga.objects.select_for_update().filter(id=carga_id).first()
    if carga is None:
        raise CargaError(f"No existe la carga {carga_id}.")
    if carga.estado != MatrizPDLCarga.BORRADOR:
        raise CargaError(
            f"La carga {carga_id} está «{carga.estado}»: solo se aplica un borrador.")
    if not carga.diff:
        raise CargaError(f"La carga {carga_id} no tiene diff. Volvé a previsualizar.")

    d = carga.diff
    hecho = {"altas": 0, "cambios": 0, "reactivaciones": 0, "retiros": 0}

    # ── sectores ──
    for nombre in d["sector"]["altas"]:
        Sector.objects.create(nombre_oficial=nombre, carga_origen_id=carga.id)
        hecho["altas"] += 1
    for nombre in d["sector"].get("reactivaciones", []):
        hecho["reactivaciones"] += Sector.objects.filter(
            nombre_oficial=nombre).update(activo=True, carga_retiro_id=None)
    for nombre in d["sector"]["retiros"]:
        hecho["retiros"] += Sector.objects.filter(nombre_oficial=nombre).update(
            activo=False, carga_retiro_id=carga.id)

    # ── objetivos ──
    for alta in d["objetivo"]["altas"]:
        ObjetivoEstrategico.objects.create(
            codigo=alta["codigo"], nombre=alta["nombre"], carga_origen_id=carga.id)
        hecho["altas"] += 1
    for cam in d["objetivo"]["cambios"]:
        hecho["cambios"] += ObjetivoEstrategico.objects.filter(
            codigo=cam["codigo"]).update(nombre=cam["a"])
    for cod in d["objetivo"].get("reactivaciones", []):
        hecho["reactivaciones"] += ObjetivoEstrategico.objects.filter(
            codigo=cod).update(activo=True, carga_retiro_id=None)
    for ret in d["objetivo"]["retiros"]:
        hecho["retiros"] += ObjetivoEstrategico.objects.filter(
            codigo=ret["codigo"]).update(activo=False, carga_retiro_id=carga.id)

    # ── programas ── (después de los objetivos: un alta puede necesitar uno nuevo)
    ids_obj = {o.codigo: o.id for o in ObjetivoEstrategico.objects.all()}
    for alta in d["programa"]["altas"]:
        objetivo_id = ids_obj.get(alta["objetivo"])
        if objetivo_id is None:
            raise CargaError(
                f"El programa {alta['codigo']} cuelga del objetivo "
                f"{alta['objetivo']}, que no existe ni lo crea esta carga.")
        ProgramaPDL.objects.create(
            codigo=alta["codigo"], nombre=alta["nombre"],
            objetivo_id=objetivo_id, carga_origen_id=carga.id)
        hecho["altas"] += 1
    for cam in d["programa"]["cambios"]:
        campos = {}
        if "nombre" in cam["campos"]:
            campos["nombre"] = cam["campos"]["nombre"]["a"]
        if "objetivo" in cam["campos"]:
            destino = ids_obj.get(cam["campos"]["objetivo"]["a"])
            if destino is None:
                raise CargaError(
                    f"El programa {cam['codigo']} se mueve al objetivo "
                    f"{cam['campos']['objetivo']['a']}, que no existe.")
            campos["objetivo_id"] = destino
        if campos:
            hecho["cambios"] += ProgramaPDL.objects.filter(
                codigo=cam["codigo"]).update(**campos)
    for cod in d["programa"].get("reactivaciones", []):
        hecho["reactivaciones"] += ProgramaPDL.objects.filter(
            codigo=cod).update(activo=True, carga_retiro_id=None)
    for ret in d["programa"]["retiros"]:
        hecho["retiros"] += ProgramaPDL.objects.filter(
            codigo=ret["codigo"]).update(activo=False, carga_retiro_id=carga.id)

    carga.estado = MatrizPDLCarga.APLICADA
    carga.aplicado_por_id = usuario_id
    carga.aplicado_at = datetime.now(timezone.utc)
    carga.save(update_fields=["estado", "aplicado_por_id", "aplicado_at"])
    return carga, hecho


def descartar(carga_id, nota=None):
    """Cierra un borrador sin aplicarlo. Queda como registro de que se miró."""
    carga = MatrizPDLCarga.objects.filter(id=carga_id).first()
    if carga is None:
        raise CargaError(f"No existe la carga {carga_id}.")
    if carga.estado != MatrizPDLCarga.BORRADOR:
        raise CargaError(f"La carga {carga_id} está «{carga.estado}».")
    carga.estado = MatrizPDLCarga.DESCARTADA
    carga.nota = nota
    carga.save(update_fields=["estado", "nota"])
    return carga
